# pyinstaller command:
# pyinstaller --noconsole --onefile --clean --strip --icon=BetterExcel.ico --optimize=2 BetterExcel.py
from pathlib import Path
import sys
import typing as t
import re
from copy import copy
from datetime import datetime
from xml.etree import ElementTree as ET
import json
import math
import secrets
import socket
import threading
import getpass
import pandas as pd
from PyQt6 import QtCore, QtGui, QtWidgets
import time
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Color
    from openpyxl.styles.colors import COLOR_INDEX
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        OPENPYXL_RICH_TEXT_AVAILABLE = True
    except Exception:
        CellRichText = None
        TextBlock = None
        InlineFont = None
        OPENPYXL_RICH_TEXT_AVAILABLE = False
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    load_workbook = None
    Font = None
    PatternFill = None
    Color = None
    CellRichText = None
    TextBlock = None
    InlineFont = None
    OPENPYXL_RICH_TEXT_AVAILABLE = False
    COLOR_INDEX = []
    OPENPYXL_AVAILABLE = False

APP_NAME = "Better Excel v1.4"
COMPACT_WINDOW_WIDTH = 470
COMPACT_WINDOW_HEIGHT = 320
TABLE_LINE_HEIGHT_PERCENT = 112
TABLE_VERTICAL_TEXT_PADDING = 10
QA_TEXT_BG_COLOR = QtGui.QColor(255, 235, 90)
QA_TEXT_FG_COLOR = QtGui.QColor(200, 0, 0)
QA_METADATA_SHEET_NAME = "__betterexcel_qa_spans__"
EDIT_METADATA_SHEET_NAME = "__betterexcel_edit_meta__"
DEFAULT_DISPLAY_FONT_FAMILY = "Inter"
DEFAULT_DISPLAY_FONT_SIZE = 10
COLLAB_LINK_SCHEME = "betterexcel"
COLLAB_PROTOCOL_VERSION = 1
COLLAB_BROADCAST_DEBOUNCE_MS = 120
COLLAB_SOCKET_TIMEOUT_SECONDS = 1.0
DROP_HINT_TEXT = "Drop a CSV / XLS / XLSX / JSONL file here\n(open file if WSL)"


def _guess_local_ipv4() -> str:
    try:
        probe = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            probe.connect(("8.8.8.8", 80))
            ip = probe.getsockname()[0]
            if ip:
                return ip
        finally:
            probe.close()
    except Exception:
        pass
    return "127.0.0.1"


def _build_collab_link(host: str, port: int, token: str) -> str:
    return f"{COLLAB_LINK_SCHEME}://{host}:{port}/{token}"


def _parse_collab_link(raw: str) -> t.Optional[tuple[str, int, str]]:
    text = (raw or "").strip()
    if not text:
        return None
    pattern = re.compile(
        rf"^(?:{COLLAB_LINK_SCHEME}://)?(?P<host>[^:/\s]+):(?P<port>\d{{1,5}})/(?P<token>[A-Za-z0-9_-]+)$"
    )
    match = pattern.match(text)
    if match is None:
        return None

    host = match.group("host").strip()
    token = match.group("token").strip()
    try:
        port = int(match.group("port"))
    except Exception:
        return None

    if not host or not token or port < 1 or port > 65535:
        return None
    return host, port, token


def _normalize_username(raw: t.Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return "Guest"
    text = re.sub(r"\s+", " ", text)
    if len(text) > 64:
        text = text[:64].rstrip()
    return text or "Guest"


def _read_local_os_username() -> str:
    try:
        return _normalize_username(getpass.getuser())
    except Exception:
        return "Guest"


def _normalize_edit_timestamp(raw: t.Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    if len(text) > 48:
        text = text[:48].rstrip()
    return text


def _now_edit_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _presence_color_for_username(username: str) -> QtGui.QColor:
    normalized = _normalize_username(username)
    hue_seed = 0
    for idx, ch in enumerate(normalized):
        hue_seed = (hue_seed + ((idx + 1) * ord(ch))) % 30
    # Keep collaborator highlights in the blue family for consistency.
    hue = 195 + int(hue_seed)
    color = QtGui.QColor.fromHsv(hue, 170, 250, 255)
    if color.isValid():
        return color
    return QtGui.QColor(65, 145, 255, 255)


def _edited_cell_color_for_username(username: str) -> QtGui.QColor:
    normalized = _normalize_username(username)
    # Stable FNV-1a hash to avoid Python's per-process randomized hash().
    seed = 2166136261
    for ch in normalized:
        seed ^= ord(ch)
        seed = (seed * 16777619) & 0xFFFFFFFF

    hue = int(seed % 360)
    sat = int(150 + ((seed >> 8) % 70))
    val = int(195 + ((seed >> 16) % 55))
    color = QtGui.QColor.fromHsv(hue, sat, val, 255)
    if color.isValid():
        return color
    return QtGui.QColor(255, 120, 0, 255)


def _json_safe_value(value: t.Any) -> t.Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    if isinstance(value, str):
        return value
    return str(value)


def _serialize_qcolor(color: t.Any) -> t.Optional[list[int]]:
    if not isinstance(color, QtGui.QColor):
        return None
    if not color.isValid():
        return None
    return [int(color.red()), int(color.green()), int(color.blue()), int(color.alpha())]


def _deserialize_qcolor(payload: t.Any) -> t.Optional[QtGui.QColor]:
    if not isinstance(payload, (list, tuple)) or len(payload) != 4:
        return None
    try:
        r, g, b, a = [max(0, min(255, int(v))) for v in payload]
    except Exception:
        return None
    color = QtGui.QColor(r, g, b, a)
    return color if color.isValid() else None


def _normalize_serialized_qa_spans(spans: t.Any) -> list[dict[str, t.Any]]:
    normalized: list[dict[str, t.Any]] = []
    for span in spans or []:
        if not isinstance(span, dict):
            continue
        try:
            start = int(span.get("start", 0))
            length = int(span.get("length", 0))
            kind = str(span.get("kind", ""))
        except Exception:
            continue
        if start < 0 or length <= 0 or kind not in ("big", "small"):
            continue
        normalized.append({"start": start, "length": length, "kind": kind})
    return normalized


def _serialize_cell_style(style: dict[str, t.Any]) -> dict[str, t.Any]:
    data: dict[str, t.Any] = {}

    qa_mark = style.get("qa_mark")
    if qa_mark in ("pass", "fail"):
        data["qa_mark"] = qa_mark

    qa_spans = _normalize_serialized_qa_spans(style.get("qa_spans", []))
    if qa_spans:
        data["qa_spans"] = qa_spans

    for color_key in ("qa_bg", "bg", "fg"):
        serialized = _serialize_qcolor(style.get(color_key))
        if serialized is not None:
            data[color_key] = serialized

    font = style.get("font")
    if isinstance(font, QtGui.QFont):
        data["font"] = font.toString()

    has_xlsx_bg = style.get("has_xlsx_bg")
    if isinstance(has_xlsx_bg, bool):
        data["has_xlsx_bg"] = has_xlsx_bg

    edited_by_raw = str(style.get("edited_by", "")).strip()
    if edited_by_raw:
        data["edited_by"] = _normalize_username(edited_by_raw)

    edited_at = _normalize_edit_timestamp(style.get("edited_at"))
    if edited_at:
        data["edited_at"] = edited_at

    return data


def _deserialize_cell_style(payload: t.Any) -> dict[str, t.Any]:
    if not isinstance(payload, dict):
        return {}

    style: dict[str, t.Any] = {}
    qa_mark = payload.get("qa_mark")
    if qa_mark in ("pass", "fail"):
        style["qa_mark"] = qa_mark

    qa_spans = _normalize_serialized_qa_spans(payload.get("qa_spans", []))
    if qa_spans:
        style["qa_spans"] = qa_spans

    for color_key in ("qa_bg", "bg", "fg"):
        restored = _deserialize_qcolor(payload.get(color_key))
        if restored is not None:
            style[color_key] = restored

    font_data = payload.get("font")
    if isinstance(font_data, str) and font_data.strip():
        font = QtGui.QFont()
        if font.fromString(font_data):
            style["font"] = font

    has_xlsx_bg = payload.get("has_xlsx_bg")
    if isinstance(has_xlsx_bg, bool):
        style["has_xlsx_bg"] = has_xlsx_bg

    edited_by_raw = str(payload.get("edited_by", "")).strip()
    if edited_by_raw:
        style["edited_by"] = _normalize_username(edited_by_raw)

    edited_at = _normalize_edit_timestamp(payload.get("edited_at"))
    if edited_at:
        style["edited_at"] = edited_at

    return style


def _serialize_cell_styles(cell_styles: dict[tuple[int, int], dict[str, t.Any]]) -> list[dict[str, t.Any]]:
    payload: list[dict[str, t.Any]] = []
    for (row, col), style in cell_styles.items():
        if not isinstance(style, dict):
            continue
        serialized_style = _serialize_cell_style(style)
        if not serialized_style:
            continue
        payload.append({"row": int(row), "col": int(col), "style": serialized_style})
    return payload


def _deserialize_cell_styles(payload: t.Any) -> dict[tuple[int, int], dict[str, t.Any]]:
    styles: dict[tuple[int, int], dict[str, t.Any]] = {}
    if not isinstance(payload, list):
        return styles

    for entry in payload:
        if not isinstance(entry, dict):
            continue
        try:
            row = int(entry.get("row"))
            col = int(entry.get("col"))
        except Exception:
            continue
        if row < 0 or col < 0:
            continue
        style = _deserialize_cell_style(entry.get("style", {}))
        if style:
            styles[(row, col)] = style
    return styles


class CollaborationServer(QtCore.QObject):
    snapshotReceived = QtCore.pyqtSignal(object)
    presenceReceived = QtCore.pyqtSignal(object)
    errorOccurred = QtCore.pyqtSignal(str)
    clientCountChanged = QtCore.pyqtSignal(int)
    guestConnected = QtCore.pyqtSignal(str, int)
    guestDisconnected = QtCore.pyqtSignal(str, int)

    def __init__(self, session_token: str, bind_host: str = "0.0.0.0", parent=None):
        super().__init__(parent)
        self.session_token = session_token
        self.bind_host = bind_host
        self.port: t.Optional[int] = None
        self._server_socket: t.Optional[socket.socket] = None
        self._stop_event = threading.Event()
        self._accept_thread: t.Optional[threading.Thread] = None
        self._clients: dict[socket.socket, tuple[threading.Lock, str]] = {}
        self._clients_lock = threading.Lock()

    @staticmethod
    def _send_json(sock: socket.socket, payload: dict[str, t.Any]):
        raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n"
        sock.sendall(raw.encode("utf-8"))

    def start(self, port: int = 0) -> bool:
        if self._server_socket is not None:
            return True
        try:
            server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            server.bind((self.bind_host, int(port)))
            server.listen(8)
            server.settimeout(COLLAB_SOCKET_TIMEOUT_SECONDS)
        except Exception as exc:
            self.errorOccurred.emit(f"Failed to start host server: {exc}")
            return False

        self._server_socket = server
        self.port = int(server.getsockname()[1])
        self._stop_event.clear()
        self._accept_thread = threading.Thread(target=self._accept_loop, daemon=True)
        self._accept_thread.start()
        return True

    def stop(self):
        self._stop_event.set()
        server = self._server_socket
        self._server_socket = None
        if server is not None:
            try:
                server.close()
            except Exception:
                pass

        with self._clients_lock:
            clients = list(self._clients.keys())
            self._clients.clear()

        for conn in clients:
            try:
                conn.shutdown(socket.SHUT_RDWR)
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

        self.clientCountChanged.emit(0)
        if self._accept_thread is not None and self._accept_thread.is_alive():
            self._accept_thread.join(timeout=0.25)
        self._accept_thread = None

    def _accept_loop(self):
        while not self._stop_event.is_set():
            if self._server_socket is None:
                break
            try:
                conn, _ = self._server_socket.accept()
            except socket.timeout:
                continue
            except OSError:
                if not self._stop_event.is_set():
                    self.errorOccurred.emit("Host listener stopped unexpectedly.")
                break
            handler = threading.Thread(target=self._client_loop, args=(conn,), daemon=True)
            handler.start()

    def _add_client(self, conn: socket.socket, username: str):
        normalized_username = _normalize_username(username)
        with self._clients_lock:
            self._clients[conn] = (threading.Lock(), normalized_username)
            client_count = len(self._clients)
        self.guestConnected.emit(normalized_username, client_count)
        self.clientCountChanged.emit(client_count)

    def _remove_client(self, conn: socket.socket):
        removed_username = "Guest"
        with self._clients_lock:
            removed = self._clients.pop(conn, None)
            existed = removed is not None
            if removed is not None:
                _, removed_username = removed
            client_count = len(self._clients)
        try:
            conn.shutdown(socket.SHUT_RDWR)
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        if existed:
            self.guestDisconnected.emit(_normalize_username(removed_username), client_count)
            self.clientCountChanged.emit(client_count)

    def _client_loop(self, conn: socket.socket):
        reader = None
        authorized = False
        try:
            conn.settimeout(8.0)
            reader = conn.makefile("r", encoding="utf-8", newline="\n")
            hello_line = reader.readline()
            if not hello_line:
                return
            try:
                hello = json.loads(hello_line)
            except Exception:
                return

            if not isinstance(hello, dict):
                return
            if hello.get("type") != "hello":
                return
            if str(hello.get("token", "")) != self.session_token:
                return

            remote_username = _normalize_username(hello.get("username"))
            self._send_json(conn, {"type": "welcome", "protocol": COLLAB_PROTOCOL_VERSION})
            conn.settimeout(None)
            authorized = True
            self._add_client(conn, remote_username)

            while not self._stop_event.is_set():
                line = reader.readline()
                if not line:
                    break
                try:
                    msg = json.loads(line)
                except Exception:
                    continue
                if not isinstance(msg, dict):
                    continue
                msg_type = msg.get("type")
                payload = msg.get("payload")
                if msg_type == "snapshot":
                    if isinstance(payload, dict):
                        self.snapshotReceived.emit(payload)
                    continue
                if msg_type == "presence":
                    if isinstance(payload, dict):
                        normalized_payload = dict(payload)
                        normalized_payload["username"] = remote_username
                        self.presenceReceived.emit(normalized_payload)
                    continue
        except Exception:
            pass
        finally:
            if reader is not None:
                try:
                    reader.close()
                except Exception:
                    pass
            if authorized:
                self._remove_client(conn)
            else:
                try:
                    conn.close()
                except Exception:
                    pass

    def broadcast_snapshot(self, snapshot: dict[str, t.Any]):
        message = {"type": "snapshot", "payload": snapshot}
        with self._clients_lock:
            clients = [(conn, lock_and_name[0]) for conn, lock_and_name in self._clients.items()]

        for conn, send_lock in clients:
            try:
                with send_lock:
                    self._send_json(conn, message)
            except Exception:
                self._remove_client(conn)

    def broadcast_presence(self, payload: dict[str, t.Any]):
        message = {"type": "presence", "payload": payload}
        with self._clients_lock:
            clients = [(conn, lock_and_name[0]) for conn, lock_and_name in self._clients.items()]

        for conn, send_lock in clients:
            try:
                with send_lock:
                    self._send_json(conn, message)
            except Exception:
                self._remove_client(conn)


class CollaborationClient(QtCore.QObject):
    snapshotReceived = QtCore.pyqtSignal(object)
    presenceReceived = QtCore.pyqtSignal(object)
    connectedChanged = QtCore.pyqtSignal(bool)
    errorOccurred = QtCore.pyqtSignal(str)

    def __init__(self, host: str, port: int, token: str, username: str = "Guest", parent=None):
        super().__init__(parent)
        self.host = host
        self.port = int(port)
        self.token = token
        self.username = _normalize_username(username)
        self._stop_event = threading.Event()
        self._thread: t.Optional[threading.Thread] = None
        self._socket: t.Optional[socket.socket] = None
        self._send_lock = threading.Lock()
        self._is_connected = False

    @staticmethod
    def _send_json(sock: socket.socket, payload: dict[str, t.Any]):
        raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n"
        sock.sendall(raw.encode("utf-8"))

    def is_connected(self) -> bool:
        return self._is_connected

    def _set_connected(self, value: bool):
        if self._is_connected != value:
            self._is_connected = value
            self.connectedChanged.emit(value)

    def start(self):
        if self._thread is not None and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_event.set()
        sock = self._socket
        self._socket = None
        if sock is not None:
            try:
                sock.shutdown(socket.SHUT_RDWR)
            except Exception:
                pass
            try:
                sock.close()
            except Exception:
                pass
        if self._thread is not None and self._thread.is_alive():
            self._thread.join(timeout=0.25)
        self._thread = None
        self._set_connected(False)

    def _run(self):
        reader = None
        sock = None
        try:
            sock = socket.create_connection((self.host, self.port), timeout=8.0)
            sock.settimeout(None)
            self._socket = sock
            self._send_json(
                sock,
                {
                    "type": "hello",
                    "token": self.token,
                    "username": self.username,
                    "protocol": COLLAB_PROTOCOL_VERSION,
                },
            )

            reader = sock.makefile("r", encoding="utf-8", newline="\n")
            welcome_line = reader.readline()
            if not welcome_line:
                self.errorOccurred.emit("Failed to connect: no handshake response from host.")
                return
            try:
                welcome = json.loads(welcome_line)
            except Exception:
                self.errorOccurred.emit("Failed to connect: invalid handshake response from host.")
                return
            if not isinstance(welcome, dict) or welcome.get("type") != "welcome":
                self.errorOccurred.emit("Failed to connect: host rejected this session.")
                return

            self._set_connected(True)

            while not self._stop_event.is_set():
                line = reader.readline()
                if not line:
                    break
                try:
                    msg = json.loads(line)
                except Exception:
                    continue
                if not isinstance(msg, dict):
                    continue
                msg_type = msg.get("type")
                payload = msg.get("payload")
                if msg_type == "snapshot":
                    if isinstance(payload, dict):
                        self.snapshotReceived.emit(payload)
                    continue
                if msg_type == "presence":
                    if isinstance(payload, dict):
                        self.presenceReceived.emit(payload)
                    continue

            if not self._stop_event.is_set():
                self.errorOccurred.emit("Disconnected from host.")
        except Exception as exc:
            if not self._stop_event.is_set():
                self.errorOccurred.emit(f"Failed to connect to host: {exc}")
        finally:
            if reader is not None:
                try:
                    reader.close()
                except Exception:
                    pass
            if sock is not None:
                try:
                    sock.close()
                except Exception:
                    pass
            self._socket = None
            self._set_connected(False)

    def send_snapshot(self, snapshot: dict[str, t.Any]) -> bool:
        sock = self._socket
        if sock is None or not self._is_connected:
            return False
        try:
            with self._send_lock:
                self._send_json(sock, {"type": "snapshot", "payload": snapshot})
            return True
        except Exception:
            return False

    def send_presence(self, payload: dict[str, t.Any]) -> bool:
        sock = self._socket
        if sock is None or not self._is_connected:
            return False
        try:
            with self._send_lock:
                self._send_json(sock, {"type": "presence", "payload": payload})
            return True
        except Exception:
            return False

# -------------------- Utility: Dark/Light palettes & styles --------------------
def apply_dark_palette(app: QtWidgets.QApplication):
    palette = QtGui.QPalette()
    palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(30, 30, 30))
    palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtCore.Qt.GlobalColor.white)
    palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(22, 22, 22))
    palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(35, 35, 35))
    palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtCore.Qt.GlobalColor.white)
    palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtCore.Qt.GlobalColor.white)
    palette.setColor(QtGui.QPalette.ColorRole.Text, QtCore.Qt.GlobalColor.white)
    palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(45, 45, 45))
    palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtCore.Qt.GlobalColor.white)
    palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtCore.Qt.GlobalColor.red)
    palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(0, 122, 204))
    palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(60, 100, 160))
    palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.white)
    app.setPalette(palette)
    app.setStyle("Fusion")


def apply_light_palette(app: QtWidgets.QApplication):
    app.setPalette(app.style().standardPalette())
    app.setStyle("Fusion")


# Subtle column zebra shading
def column_shade_color(is_dark: bool, index: int) -> QtGui.QColor:
    """FINAL VERSION – subtle tinted whites"""
    cycle_len = 12
    step = index % cycle_len
    hue = int((step / cycle_len) * 360)
    saturation = 100  # visible but not garish
    lightness = 30
    return QtGui.QColor.fromHsl(hue, saturation, lightness)


def _qcolor_from_argb_hex(value: str) -> t.Optional[QtGui.QColor]:
    if not value:
        return None
    val = value.strip().lstrip("#")
    if len(val) == 8:
        a = int(val[0:2], 16)
        if a == 0:
            return None
        r = int(val[2:4], 16)
        g = int(val[4:6], 16)
        b = int(val[6:8], 16)
        color = QtGui.QColor(r, g, b)
        color.setAlpha(a)
        return color
    if len(val) == 6:
        return QtGui.QColor(f"#{val}")
    return None


def _argb_hex_from_qcolor(color: QtGui.QColor) -> str:
    a = max(0, min(255, color.alpha()))
    r = max(0, min(255, color.red()))
    g = max(0, min(255, color.green()))
    b = max(0, min(255, color.blue()))
    return f"{a:02X}{r:02X}{g:02X}{b:02X}"


def _apply_excel_tint(color: QtGui.QColor, tint: t.Optional[float]) -> QtGui.QColor:
    if tint is None:
        return color
    try:
        tint_value = float(tint)
    except Exception:
        return color
    if abs(tint_value) < 1e-9:
        return color

    def _apply_channel(ch: int) -> int:
        if tint_value < 0:
            return max(0, min(255, int(ch * (1.0 + tint_value))))
        return max(0, min(255, int(ch * (1.0 - tint_value) + 255.0 * tint_value)))

    tinted = QtGui.QColor(_apply_channel(color.red()), _apply_channel(color.green()), _apply_channel(color.blue()))
    tinted.setAlpha(color.alpha())
    return tinted


def _extract_excel_theme_palette(workbook) -> dict[int, QtGui.QColor]:
    palette: dict[int, QtGui.QColor] = {}
    if not workbook:
        return palette

    raw_theme = getattr(workbook, "loaded_theme", None)
    if not raw_theme:
        return palette

    if isinstance(raw_theme, bytes):
        raw_theme = raw_theme.decode("utf-8", errors="ignore")

    try:
        root = ET.fromstring(raw_theme)
    except Exception:
        return palette

    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    clr_scheme = root.find(".//a:themeElements/a:clrScheme", ns)
    if clr_scheme is None:
        return palette

    scheme_names = [
        "lt1",
        "dk1",
        "lt2",
        "dk2",
        "accent1",
        "accent2",
        "accent3",
        "accent4",
        "accent5",
        "accent6",
        "hlink",
        "folHlink",
    ]

    for idx, name in enumerate(scheme_names):
        color_node = clr_scheme.find(f"a:{name}", ns)
        if color_node is None:
            continue

        srgb = color_node.find("a:srgbClr", ns)
        if srgb is not None and srgb.get("val"):
            candidate = QtGui.QColor(f"#{srgb.get('val')}")
            if candidate.isValid():
                palette[idx] = candidate
            continue

        sys_clr = color_node.find("a:sysClr", ns)
        if sys_clr is not None:
            fallback = sys_clr.get("lastClr") or sys_clr.get("val")
            if fallback:
                candidate = QtGui.QColor(f"#{fallback}")
                if candidate.isValid():
                    palette[idx] = candidate

    return palette


def _qcolor_from_openpyxl_color(color, theme_palette: t.Optional[dict[int, QtGui.QColor]] = None) -> t.Optional[QtGui.QColor]:
    if color is None:
        return None

    if getattr(color, "type", None) == "rgb" and getattr(color, "rgb", None):
        return _qcolor_from_argb_hex(color.rgb)

    if getattr(color, "type", None) == "indexed" and getattr(color, "indexed", None) is not None:
        try:
            idx = int(color.indexed)
        except Exception:
            return None
        if 0 <= idx < len(COLOR_INDEX):
            return _qcolor_from_argb_hex(COLOR_INDEX[idx])

    if getattr(color, "type", None) == "theme" and getattr(color, "theme", None) is not None:
        if not theme_palette:
            return None
        try:
            theme_idx = int(color.theme)
        except Exception:
            return None
        base = theme_palette.get(theme_idx)
        if base is None:
            return None
        tinted = _apply_excel_tint(base, getattr(color, "tint", None))
        return tinted if tinted.isValid() else None

    return None


def _is_default_excel_font_color(color) -> bool:
    if color is None:
        return True

    if getattr(color, "type", None) != "theme" or getattr(color, "theme", None) is None:
        return False

    try:
        theme_idx = int(color.theme)
    except Exception:
        return False

    tint = getattr(color, "tint", None)
    try:
        tint_is_zero = tint is None or abs(float(tint)) < 1e-9
    except Exception:
        tint_is_zero = False

    # Excel's default text color in the common Office theme is Dark1 (theme index 1).
    return theme_idx == 1 and tint_is_zero


def _is_effectively_default_excel_font(font) -> bool:
    if font is None or not OPENPYXL_AVAILABLE:
        return True

    default_font = DEFAULT_XLSX_FONT
    if default_font is None:
        return False

    def _norm_opt(value):
        return None if value is None else str(value)

    def _norm_underline(value):
        if value is None:
            return None
        val = str(value).strip().lower()
        return None if not val or val == "none" else val

    def _norm_size(value, fallback):
        if value is None:
            return fallback
        try:
            return float(value)
        except Exception:
            return fallback

    default_name = _norm_opt(getattr(default_font, "name", None))
    default_size = _norm_size(getattr(default_font, "sz", None), 11.0)

    name = _norm_opt(getattr(font, "name", None))
    if name is not None and default_name is not None and name != default_name:
        return False

    size = _norm_size(getattr(font, "sz", None), default_size)
    if abs(size - default_size) > 1e-6:
        return False

    if bool(getattr(font, "bold", False)):
        return False
    if bool(getattr(font, "italic", False)):
        return False
    if bool(getattr(font, "strike", False)):
        return False

    if _norm_underline(getattr(font, "underline", None)) != _norm_underline(getattr(default_font, "underline", None)):
        return False

    if _norm_opt(getattr(font, "vertAlign", None)) != _norm_opt(getattr(default_font, "vertAlign", None)):
        return False

    if _norm_opt(getattr(font, "charset", None)) not in (None, _norm_opt(getattr(default_font, "charset", None))):
        return False
    if _norm_opt(getattr(font, "family", None)) not in (None, _norm_opt(getattr(default_font, "family", None))):
        return False
    if _norm_opt(getattr(font, "scheme", None)) not in (None, _norm_opt(getattr(default_font, "scheme", None))):
        return False

    return _is_default_excel_font_color(getattr(font, "color", None))


def _color_luminance(color: QtGui.QColor) -> float:
    return 0.2126 * color.red() + 0.7152 * color.green() + 0.0722 * color.blue()


def _colors_close(a: QtGui.QColor, b: QtGui.QColor, tol: int = 8) -> bool:
    return (
        abs(a.red() - b.red()) <= tol
        and abs(a.green() - b.green()) <= tol
        and abs(a.blue() - b.blue()) <= tol
        and abs(a.alpha() - b.alpha()) <= tol
    )


def _normalize_qa_spans(
    spans: t.Optional[list[dict[str, t.Any]]],
    text_len: int,
) -> list[dict[str, t.Any]]:
    max_len = max(0, int(text_len))
    clean: list[dict[str, t.Any]] = []
    for span in spans or []:
        try:
            start = int(span.get("start", 0))
            length = int(span.get("length", 0))
            kind = str(span.get("kind", ""))
        except Exception:
            continue
        if kind not in ("big", "small") or length <= 0:
            continue
        start = max(0, start)
        if start >= max_len:
            continue
        end = min(max_len, start + length)
        if end <= start:
            continue
        clean.append({"start": start, "length": end - start, "kind": kind})

    if not clean:
        return []

    ordered = sorted(clean, key=lambda s: (int(s["start"]), int(s["length"])))
    merged: list[dict[str, t.Any]] = [ordered[0]]
    for span in ordered[1:]:
        prev = merged[-1]
        prev_end = int(prev["start"]) + int(prev["length"])
        cur_start = int(span["start"])
        cur_end = cur_start + int(span["length"])
        if span["kind"] == prev["kind"] and cur_start <= prev_end:
            prev["length"] = max(prev_end, cur_end) - int(prev["start"])
        else:
            merged.append(span)
    return merged


def _normalize_text_and_qa_spans_for_rich_text(
    text: str,
    spans: t.Optional[list[dict[str, t.Any]]],
) -> tuple[str, list[dict[str, t.Any]]]:
    src = text or ""
    if not src:
        return "", []

    normalized_spans = _normalize_qa_spans(spans, len(src))
    if not any(ch in src for ch in ("\r", "\u2028", "\u2029")):
        return src, normalized_spans

    boundary_map = [0] * (len(src) + 1)
    out_chars: list[str] = []
    i = 0
    j = 0
    while i < len(src):
        boundary_map[i] = j
        ch = src[i]

        if ch == "\r":
            if i + 1 < len(src) and src[i + 1] == "\n":
                boundary_map[i + 1] = j
                out_chars.append("\n")
                j += 1
                i += 2
                boundary_map[i] = j
                continue
            out_chars.append("\n")
            j += 1
            i += 1
            boundary_map[i] = j
            continue

        if ch == "\n":
            if i + 1 < len(src) and src[i + 1] == "\r":
                boundary_map[i + 1] = j
                out_chars.append("\n")
                j += 1
                i += 2
                boundary_map[i] = j
                continue
            out_chars.append("\n")
            j += 1
            i += 1
            boundary_map[i] = j
            continue

        if ch in ("\u2028", "\u2029"):
            out_chars.append("\n")
            j += 1
            i += 1
            boundary_map[i] = j
            continue

        out_chars.append(ch)
        j += 1
        i += 1
        boundary_map[i] = j

    normalized_text = "".join(out_chars)
    remapped_spans: list[dict[str, t.Any]] = []
    for span in normalized_spans:
        start = max(0, min(len(src), int(span["start"])))
        end = max(start, min(len(src), start + int(span["length"])))
        kind = str(span["kind"])
        remapped_start = boundary_map[start]
        remapped_end = boundary_map[end]
        if remapped_end <= remapped_start:
            continue
        remapped_spans.append(
            {"start": remapped_start, "length": remapped_end - remapped_start, "kind": kind}
        )

    return normalized_text, _normalize_qa_spans(remapped_spans, len(normalized_text))


def _load_workbook_with_rich_text(path: Path):
    if OPENPYXL_RICH_TEXT_AVAILABLE:
        try:
            return load_workbook(path, data_only=False, rich_text=True)
        except TypeError:
            pass
    return load_workbook(path, data_only=False)


def _qa_kind_from_openpyxl_font(font, theme_palette: t.Optional[dict[int, QtGui.QColor]] = None) -> t.Optional[str]:
    if font is None:
        return None

    color_value = getattr(font, "color", None)
    if isinstance(color_value, str):
        qcolor = _qcolor_from_argb_hex(color_value)
    else:
        qcolor = _qcolor_from_openpyxl_color(color_value, theme_palette)
    if qcolor is None or not _colors_close(qcolor, QA_TEXT_FG_COLOR, tol=20):
        return None

    is_bold = bool(getattr(font, "b", False))
    is_italic = bool(getattr(font, "i", False))
    if is_bold and not is_italic:
        return "big"
    if is_italic and not is_bold:
        return "small"
    return None


def _extract_qa_spans_from_rich_text(
    rich_value,
    theme_palette: t.Optional[dict[int, QtGui.QColor]] = None,
) -> list[dict[str, t.Any]]:
    if not OPENPYXL_RICH_TEXT_AVAILABLE or CellRichText is None or TextBlock is None:
        return []
    if not isinstance(rich_value, CellRichText):
        return []

    spans: list[dict[str, t.Any]] = []
    cursor = 0
    for part in rich_value:
        if isinstance(part, str):
            cursor += len(part)
            continue

        if isinstance(part, TextBlock):
            text = part.text or ""
            kind = _qa_kind_from_openpyxl_font(part.font, theme_palette)
            if kind and text:
                spans.append({"start": cursor, "length": len(text), "kind": kind})
            cursor += len(text)
            continue

        raw_text = str(getattr(part, "text", "") or "")
        cursor += len(raw_text)

    return _normalize_qa_spans(spans, cursor)


def _build_inline_font_for_qa(base_font, kind: str):
    if not OPENPYXL_RICH_TEXT_AVAILABLE or InlineFont is None:
        return None
    if kind not in ("big", "small"):
        return None

    run_font = InlineFont()
    if base_font is not None:
        family = getattr(base_font, "name", None)
        size = getattr(base_font, "sz", None)
        charset = getattr(base_font, "charset", None)
        family_id = getattr(base_font, "family", None)
        underline = getattr(base_font, "u", None)
        scheme = getattr(base_font, "scheme", None)

        if family:
            run_font.rFont = str(family)
        if size:
            try:
                run_font.sz = float(size)
            except Exception:
                pass
        if charset is not None:
            run_font.charset = charset
        if family_id is not None:
            run_font.family = family_id
        if underline:
            run_font.u = underline
        if scheme:
            run_font.scheme = scheme

    qa_color = _argb_hex_from_qcolor(QA_TEXT_FG_COLOR)
    try:
        run_font.color = Color(rgb=qa_color) if Color is not None else qa_color
    except Exception:
        run_font.color = qa_color

    run_font.b = kind == "big"
    run_font.i = kind == "small"
    return run_font


def _build_rich_text_value_for_cell(text: str, qa_spans: list[dict[str, t.Any]], base_font):
    safe_text, normalized_spans = _normalize_text_and_qa_spans_for_rich_text(text or "", qa_spans)
    if not OPENPYXL_RICH_TEXT_AVAILABLE or CellRichText is None or TextBlock is None:
        return safe_text

    if not normalized_spans:
        return safe_text

    rich_value = CellRichText()
    cursor = 0
    for span in normalized_spans:
        start = int(span["start"])
        end = start + int(span["length"])
        kind = str(span["kind"])

        if cursor < start:
            rich_value.append(safe_text[cursor:start])

        chunk = safe_text[start:end]
        if chunk:
            inline_font = _build_inline_font_for_qa(base_font, kind)
            if inline_font is not None:
                try:
                    rich_value.append(TextBlock(inline_font, chunk))
                except Exception:
                    rich_value.append(chunk)
            else:
                rich_value.append(chunk)
        cursor = end

    if cursor < len(safe_text):
        rich_value.append(safe_text[cursor:])

    return rich_value


def _collect_qa_spans_metadata_rows(
    sheet_name: str,
    df: pd.DataFrame,
    cell_styles: dict[tuple[int, int], dict[str, t.Any]],
) -> list[tuple[str, int, int, int, int, str]]:
    rows: list[tuple[str, int, int, int, int, str]] = []
    for (row_idx, col_idx), style in cell_styles.items():
        qa_spans = style.get("qa_spans", [])
        if not qa_spans:
            continue
        if row_idx < 0 or col_idx < 0 or row_idx >= df.shape[0] or col_idx >= df.shape[1]:
            continue
        value = df.iat[row_idx, col_idx]
        text_value = "" if pd.isna(value) else str(value)
        _, normalized_spans = _normalize_text_and_qa_spans_for_rich_text(text_value, qa_spans)
        for span in normalized_spans:
            rows.append(
                (
                    sheet_name,
                    int(row_idx),
                    int(col_idx),
                    int(span["start"]),
                    int(span["length"]),
                    str(span["kind"]),
                )
            )
    return rows


def _write_qa_spans_metadata_sheet(
    wb,
    sheet_name: str,
    df: pd.DataFrame,
    cell_styles: dict[tuple[int, int], dict[str, t.Any]],
):
    if QA_METADATA_SHEET_NAME in wb.sheetnames:
        stale = wb[QA_METADATA_SHEET_NAME]
        wb.remove(stale)

    rows = _collect_qa_spans_metadata_rows(sheet_name, df, cell_styles)
    if not rows:
        return

    meta_ws = wb.create_sheet(QA_METADATA_SHEET_NAME)
    meta_ws.sheet_state = "veryHidden"
    meta_ws.append(["sheet", "row", "col", "start", "length", "kind"])
    for row in rows:
        meta_ws.append(list(row))


def _read_qa_spans_metadata_sheet(wb, target_sheet_name: str) -> dict[tuple[int, int], list[dict[str, t.Any]]]:
    result: dict[tuple[int, int], list[dict[str, t.Any]]] = {}
    if QA_METADATA_SHEET_NAME not in wb.sheetnames:
        return result

    meta_ws = wb[QA_METADATA_SHEET_NAME]
    for values in meta_ws.iter_rows(min_row=2, values_only=True):
        if not values or len(values) < 6:
            continue
        sheet_name, row_idx, col_idx, start, length, kind = values[:6]
        if str(sheet_name) != str(target_sheet_name):
            continue
        try:
            r = int(row_idx)
            c = int(col_idx)
            s = int(start)
            l = int(length)
            k = str(kind)
        except Exception:
            continue
        if r < 0 or c < 0 or s < 0 or l <= 0 or k not in ("big", "small"):
            continue
        result.setdefault((r, c), []).append({"start": s, "length": l, "kind": k})

    return result


def _collect_edit_metadata_rows(
    sheet_name: str,
    df: pd.DataFrame,
    cell_styles: dict[tuple[int, int], dict[str, t.Any]],
) -> list[tuple[str, int, int, str, str]]:
    rows: list[tuple[str, int, int, str, str]] = []
    for (row_idx, col_idx), style in cell_styles.items():
        if row_idx < 0 or col_idx < 0 or row_idx >= df.shape[0] or col_idx >= df.shape[1]:
            continue
        edited_by_raw = str(style.get("edited_by", "")).strip()
        edited_at = _normalize_edit_timestamp(style.get("edited_at"))
        if not edited_by_raw or not edited_at:
            continue
        rows.append(
            (
                str(sheet_name),
                int(row_idx),
                int(col_idx),
                _normalize_username(edited_by_raw),
                edited_at,
            )
        )
    return rows


def _write_edit_metadata_sheet(
    wb,
    sheet_name: str,
    df: pd.DataFrame,
    cell_styles: dict[tuple[int, int], dict[str, t.Any]],
):
    if EDIT_METADATA_SHEET_NAME in wb.sheetnames:
        stale = wb[EDIT_METADATA_SHEET_NAME]
        wb.remove(stale)

    rows = _collect_edit_metadata_rows(sheet_name, df, cell_styles)
    if not rows:
        return

    meta_ws = wb.create_sheet(EDIT_METADATA_SHEET_NAME)
    meta_ws.sheet_state = "veryHidden"
    meta_ws.append(["sheet", "row", "col", "edited_by", "edited_at"])
    for row in rows:
        meta_ws.append(list(row))


def _read_edit_metadata_sheet(wb, target_sheet_name: str) -> dict[tuple[int, int], dict[str, str]]:
    result: dict[tuple[int, int], dict[str, str]] = {}
    if EDIT_METADATA_SHEET_NAME not in wb.sheetnames:
        return result

    meta_ws = wb[EDIT_METADATA_SHEET_NAME]
    for values in meta_ws.iter_rows(min_row=2, values_only=True):
        if not values or len(values) < 5:
            continue
        sheet_name, row_idx, col_idx, edited_by, edited_at = values[:5]
        if str(sheet_name) != str(target_sheet_name):
            continue
        try:
            r = int(row_idx)
            c = int(col_idx)
        except Exception:
            continue
        if r < 0 or c < 0:
            continue

        edited_by_raw = str(edited_by or "").strip()
        edited_at_text = _normalize_edit_timestamp(edited_at)
        if not edited_by_raw or not edited_at_text:
            continue

        result[(r, c)] = {
            "edited_by": _normalize_username(edited_by_raw),
            "edited_at": edited_at_text,
        }
    return result


def _is_effectively_white(color: t.Optional[QtGui.QColor], threshold: int = 242) -> bool:
    if color is None or not color.isValid() or color.alpha() == 0:
        return True
    return color.red() >= threshold and color.green() >= threshold and color.blue() >= threshold


def _invert_black_white_display_color(color: QtGui.QColor) -> QtGui.QColor:
    if not color.isValid():
        return color
    if color.alpha() == 0:
        return color

    # Display-only swap: dark -> white, light -> black.
    if _color_luminance(color) <= 24:
        inverted = QtGui.QColor(255, 255, 255)
        inverted.setAlpha(color.alpha())
        return inverted
    if _color_luminance(color) >= 231:
        inverted = QtGui.QColor(0, 0, 0)
        inverted.setAlpha(color.alpha())
        return inverted
    return color


def _apply_document_line_spacing(doc: QtGui.QTextDocument):
    doc.setDocumentMargin(1.0)
    cursor = QtGui.QTextCursor(doc)
    cursor.select(QtGui.QTextCursor.SelectionType.Document)
    block_format = QtGui.QTextBlockFormat()
    line_height_type = int(QtGui.QTextBlockFormat.LineHeightTypes.ProportionalHeight.value)
    block_format.setLineHeight(TABLE_LINE_HEIGHT_PERCENT, line_height_type)
    cursor.mergeBlockFormat(block_format)


DEFAULT_XLSX_FONT = Font() if OPENPYXL_AVAILABLE else None
DEFAULT_XLSX_FILL = PatternFill() if OPENPYXL_AVAILABLE else None
DEFAULT_EXPORT_XLSX_FONT = (
    Font(name=DEFAULT_DISPLAY_FONT_FAMILY, sz=DEFAULT_DISPLAY_FONT_SIZE)
    if OPENPYXL_AVAILABLE
    else None
)


def load_xlsx_dataframe_with_styles(
    path: Path,
    sheet_name: t.Optional[str] = None,
) -> tuple[pd.DataFrame, dict[tuple[int, int], dict[str, t.Any]], str, list[str], t.Any]:
    if not OPENPYXL_AVAILABLE:
        target = sheet_name if sheet_name else 0
        df = pd.read_excel(path, sheet_name=target)
        resolved_name = sheet_name or "Sheet1"
        return df, {}, resolved_name, [resolved_name], None

    wb = _load_workbook_with_rich_text(path)
    hidden_meta_sheets = {QA_METADATA_SHEET_NAME, EDIT_METADATA_SHEET_NAME}
    sheet_names = [name for name in wb.sheetnames if name not in hidden_meta_sheets]
    if not sheet_names:
        return pd.DataFrame(), {}, "Sheet1", [], None

    workbook_base_font = None
    workbook_fonts = getattr(wb, "_fonts", None)
    if workbook_fonts:
        try:
            workbook_base_font = copy(workbook_fonts[0])
        except Exception:
            workbook_base_font = None
    if workbook_base_font is None and DEFAULT_XLSX_FONT is not None:
        workbook_base_font = copy(DEFAULT_XLSX_FONT)

    target_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
    ws = wb[target_sheet]
    df = pd.read_excel(path, engine="openpyxl", sheet_name=target_sheet)
    theme_palette = _extract_excel_theme_palette(wb)
    qa_spans_metadata = _read_qa_spans_metadata_sheet(wb, target_sheet)
    edit_metadata = _read_edit_metadata_sheet(wb, target_sheet)

    cell_styles: dict[tuple[int, int], dict[str, t.Any]] = {}
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell = ws.cell(row=r + 2, column=c + 1)
            style: dict[str, t.Any] = {}
            cell_font = copy(cell.font) if getattr(cell, "font", None) is not None else None

            fill_is_custom = bool(cell.fill and cell.fill != DEFAULT_XLSX_FILL and cell.fill.fill_type)
            font_is_custom = bool(cell_font) and not _is_effectively_default_excel_font(cell_font)

            if fill_is_custom:
                bg = (
                    _qcolor_from_openpyxl_color(cell.fill.fgColor, theme_palette)
                    or _qcolor_from_openpyxl_color(cell.fill.bgColor, theme_palette)
                    or _qcolor_from_openpyxl_color(getattr(cell.fill, "start_color", None), theme_palette)
                )
                if bg is not None and not _is_effectively_white(bg):
                    style["has_xlsx_bg"] = True
                    style["bg"] = bg
                style["xlsx_fill"] = copy(cell.fill)

            if font_is_custom:
                qfont = QtGui.QFont(QtWidgets.QApplication.font())
                if cell_font.name:
                    qfont.setFamily(str(cell_font.name))
                if cell_font.sz:
                    qfont.setPointSizeF(float(cell_font.sz))
                qfont.setBold(bool(cell_font.bold))
                qfont.setItalic(bool(cell_font.italic))
                qfont.setUnderline(bool(cell_font.underline and cell_font.underline != "none"))
                qfont.setStrikeOut(bool(cell_font.strike))
                style["font"] = qfont

                if not _is_default_excel_font_color(cell_font.color):
                    fg = _qcolor_from_openpyxl_color(cell_font.color, theme_palette)
                    if fg is not None:
                        style["fg"] = fg
                style["xlsx_font"] = cell_font

            qa_spans = qa_spans_metadata.get((r, c), [])
            if qa_spans:
                cell_value = df.iat[r, c]
                text_value = "" if pd.isna(cell_value) else str(cell_value)
                _, qa_spans = _normalize_text_and_qa_spans_for_rich_text(text_value, qa_spans)
            else:
                qa_spans = _extract_qa_spans_from_rich_text(cell.value, theme_palette)
            if qa_spans:
                style["qa_spans"] = qa_spans

            edit_meta = edit_metadata.get((r, c))
            if edit_meta:
                edited_by_raw = str(edit_meta.get("edited_by", "")).strip()
                edited_at = _normalize_edit_timestamp(edit_meta.get("edited_at"))
                if edited_by_raw:
                    style["edited_by"] = _normalize_username(edited_by_raw)
                if edited_at:
                    style["edited_at"] = edited_at

            if style:
                cell_styles[(r, c)] = style

    return df, cell_styles, ws.title, sheet_names, workbook_base_font


def write_xlsx_with_styles(
    path: Path,
    df: pd.DataFrame,
    cell_styles: dict[tuple[int, int], dict[str, t.Any]],
    sheet_name: str = "Sheet1",
    base_font = None,
):
    if not OPENPYXL_AVAILABLE:
        df.to_excel(path, index=False)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"
    effective_base_font = copy(base_font) if base_font is not None else None

    for c, col_name in enumerate(df.columns, start=1):
        header_cell = ws.cell(row=1, column=c, value=str(col_name))
        if effective_base_font is not None:
            header_cell.font = copy(effective_base_font)

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            value = df.iat[r, c]
            if pd.isna(value):
                value = None
            elif isinstance(value, str):
                value = (
                    value.replace("\r\n", "\n")
                    .replace("\n\r", "\n")
                    .replace("\r", "\n")
                    .replace("\u2028", "\n")
                    .replace("\u2029", "\n")
                )
            cell = ws.cell(row=r + 2, column=c + 1, value=value)
            if effective_base_font is not None:
                cell.font = copy(effective_base_font)

            style = cell_styles.get((r, c))
            if not style:
                continue

            if style.get("xlsx_font") is not None:
                cell.font = copy(style["xlsx_font"])
            qa_bg = style.get("qa_bg")
            if isinstance(qa_bg, QtGui.QColor) and qa_bg.isValid():
                cell.fill = PatternFill(fill_type="solid", fgColor=_argb_hex_from_qcolor(qa_bg))
            elif style.get("xlsx_fill") is not None:
                cell.fill = copy(style["xlsx_fill"])

            qa_spans = style.get("qa_spans", [])
            if qa_spans and value is not None:
                text_value = str(value)
                normalized_text, _ = _normalize_text_and_qa_spans_for_rich_text(text_value, qa_spans)
                # Rich-text runs can introduce display line-break artifacts in some Excel cells.
                # Keep multiline cells plain and rely on metadata for restoring QA spans.
                if "\n" in normalized_text:
                    cell.value = normalized_text
                else:
                    cell.value = _build_rich_text_value_for_cell(text_value, qa_spans, cell.font)

    _write_qa_spans_metadata_sheet(wb, ws.title, df, cell_styles)
    _write_edit_metadata_sheet(wb, ws.title, df, cell_styles)

    wb.save(path)


# -------------------- Data Model for DataFrame --------------------
class DataFrameModel(QtCore.QAbstractTableModel):
    dataChangedHard = QtCore.pyqtSignal()     
    matchesChanged = QtCore.pyqtSignal(int)    
    dirtyStateChanged = QtCore.pyqtSignal(bool)
    def __init__(
        self,
        df: pd.DataFrame,
        cell_styles: t.Optional[dict[tuple[int, int], dict[str, t.Any]]] = None,
        source_format: str = "",
        xlsx_sheet_name: str = "Sheet1",
        xlsx_base_font = None,
    ):
        super().__init__()
        self._df = df
        self.search_term: str = ""
        self.replace_term: str = ""
        self._is_dark = True
        self._matches: list[QtCore.QModelIndex] = []
        self.case_sensitive: bool = False
        self._is_dirty = False
        self._cell_styles: dict[tuple[int, int], dict[str, t.Any]] = cell_styles or {}
        self._edit_actor = "Guest"
        self.source_format = source_format
        self.xlsx_sheet_name = xlsx_sheet_name
        self.xlsx_base_font = xlsx_base_font
        
    @property
    def df(self) -> pd.DataFrame:
        return self._df

    @df.setter
    def df(self, value: pd.DataFrame):
        self._df = value

    @property
    def is_dirty(self) -> bool:
        return self._is_dirty

    @is_dirty.setter
    def is_dirty(self, value: bool):
        """Automatically emit signal when dirty state changes"""
        if self._is_dirty != value:
            self._is_dirty = value
            self.dirtyStateChanged.emit(value)

    def rowCount(self, parent=None):
        return self.data_row_count() + 1

    def columnCount(self, parent=None):
        return self.data_column_count() + 1

    def data_row_count(self) -> int:
        return len(self._df.index)

    def data_column_count(self) -> int:
        return len(self._df.columns)

    def has_cell_styles(self) -> bool:
        return bool(self._cell_styles)

    def cell_styles(self) -> dict[tuple[int, int], dict[str, t.Any]]:
        return self._cell_styles

    def _cell_style(self, row: int, col: int) -> dict[str, t.Any]:
        return self._cell_styles.get((row, col), {})

    def set_edit_actor(self, username: str):
        self._edit_actor = _normalize_username(username)

    def _stamp_cell_edit_metadata(
        self,
        row: int,
        col: int,
        *,
        edited_by: t.Optional[str] = None,
        edited_at: t.Optional[str] = None,
    ):
        key = (row, col)
        style = dict(self._cell_styles.get(key, {}))
        actor_raw = str(edited_by or "").strip()
        actor = _normalize_username(actor_raw) if actor_raw else _normalize_username(self._edit_actor)
        when = _normalize_edit_timestamp(edited_at)
        if not when:
            when = _now_edit_timestamp()
        style["edited_by"] = actor
        style["edited_at"] = when
        self._cell_styles[key] = style

    def cell_edit_metadata(self, row: int, col: int) -> t.Optional[tuple[str, str]]:
        if row < 0 or col < 0:
            return None
        style = self._cell_style(row, col)
        edited_by_raw = str(style.get("edited_by", "")).strip()
        edited_at = _normalize_edit_timestamp(style.get("edited_at"))
        if not edited_by_raw or not edited_at:
            return None
        return _normalize_username(edited_by_raw), edited_at

    def reorder_cell_styles(self, visual_order: list[int]):
        if not self._cell_styles:
            return
        col_map = {old_col: new_col for new_col, old_col in enumerate(visual_order)}
        new_styles: dict[tuple[int, int], dict[str, t.Any]] = {}
        for (row, col), style in self._cell_styles.items():
            if col in col_map:
                new_styles[(row, col_map[col])] = style
        self._cell_styles = new_styles

    def drop_style_columns(self, col_indices: list[int]):
        if not self._cell_styles:
            return
        to_drop = sorted(set(col_indices))
        to_drop_set = set(to_drop)
        new_styles: dict[tuple[int, int], dict[str, t.Any]] = {}
        for (row, col), style in self._cell_styles.items():
            if col in to_drop_set:
                continue
            shift = sum(1 for idx in to_drop if idx < col)
            new_styles[(row, col - shift)] = style
        self._cell_styles = new_styles

    def drop_style_rows(self, row_indices: list[int]):
        if not self._cell_styles:
            return
        to_drop = sorted(set(row_indices))
        to_drop_set = set(to_drop)
        new_styles: dict[tuple[int, int], dict[str, t.Any]] = {}
        for (row, col), style in self._cell_styles.items():
            if row in to_drop_set:
                continue
            shift = sum(1 for idx in to_drop if idx < row)
            new_styles[(row - shift, col)] = style
        self._cell_styles = new_styles

    def set_cell_qa_mark(
        self,
        row: int,
        col: int,
        mark: t.Optional[str],
        *,
        edited_by: t.Optional[str] = None,
        edited_at: t.Optional[str] = None,
    ) -> bool:
        if row < 0 or col < 0 or row >= self.data_row_count() or col >= self.data_column_count():
            return False

        key = (row, col)
        style = dict(self._cell_styles.get(key, {}))
        old_mark = style.get("qa_mark")
        target_mark = mark if mark in ("pass", "fail") else None
        if old_mark == target_mark:
            return True

        if target_mark == "pass":
            style["qa_mark"] = "pass"
            style["qa_bg"] = QtGui.QColor(28, 120, 50)
        elif target_mark == "fail":
            style["qa_mark"] = "fail"
            style["qa_bg"] = QtGui.QColor(145, 35, 35)
        else:
            style.pop("qa_mark", None)
            style.pop("qa_bg", None)

        actor_raw = str(edited_by or "").strip()
        actor = _normalize_username(actor_raw) if actor_raw else _normalize_username(self._edit_actor)
        when = _normalize_edit_timestamp(edited_at)
        if not when:
            when = _now_edit_timestamp()
        style["edited_by"] = actor
        style["edited_at"] = when

        if style:
            self._cell_styles[key] = style
        elif key in self._cell_styles:
            del self._cell_styles[key]

        idx = self.index(row, col)
        self.dataChanged.emit(idx, idx, [QtCore.Qt.ItemDataRole.BackgroundRole])
        self.is_dirty = True
        return True

    def cell_qa_spans(self, row: int, col: int) -> list[dict[str, t.Any]]:
        spans = self._cell_style(row, col).get("qa_spans", [])
        if isinstance(spans, list):
            return spans
        return []

    def set_cell_qa_spans(
        self,
        row: int,
        col: int,
        spans: list[dict[str, t.Any]],
        emit_data_changed: bool = True,
        *,
        edited_by: t.Optional[str] = None,
        edited_at: t.Optional[str] = None,
    ) -> bool:
        if row < 0 or col < 0 or row >= self.data_row_count() or col >= self.data_column_count():
            return False

        normalized: list[dict[str, t.Any]] = []
        for span in spans or []:
            try:
                start = int(span.get("start", 0))
                length = int(span.get("length", 0))
                kind = str(span.get("kind", ""))
            except Exception:
                continue
            if start < 0 or length <= 0 or kind not in ("big", "small"):
                continue
            normalized.append({"start": start, "length": length, "kind": kind})

        key = (row, col)
        style = dict(self._cell_styles.get(key, {}))
        old_spans = style.get("qa_spans", [])
        if old_spans == normalized:
            return True

        if normalized:
            style["qa_spans"] = normalized
        else:
            style.pop("qa_spans", None)

        actor_raw = str(edited_by or "").strip()
        actor = _normalize_username(actor_raw) if actor_raw else _normalize_username(self._edit_actor)
        when = _normalize_edit_timestamp(edited_at)
        if not when:
            when = _now_edit_timestamp()
        style["edited_by"] = actor
        style["edited_at"] = when

        if style:
            self._cell_styles[key] = style
        elif key in self._cell_styles:
            del self._cell_styles[key]

        if emit_data_changed:
            idx = self.index(row, col)
            self.dataChanged.emit(idx, idx, [QtCore.Qt.ItemDataRole.DisplayRole])
        self.is_dirty = True
        return True

    def data(self, index, role=QtCore.Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        r, c = index.row(), index.column()
        
        # Handle the "+" row/column sentinel cells.
        if r == self.data_row_count() or c == self.data_column_count():
            if role == QtCore.Qt.ItemDataRole.DisplayRole:
                return ""
            elif role == QtCore.Qt.ItemDataRole.BackgroundRole:
                return QtGui.QBrush(QtGui.QColor(60, 60, 60))
            return None
        
        # Existing logic for regular columns
        val = self._df.iat[r, c]
        if role in (QtCore.Qt.ItemDataRole.DisplayRole, QtCore.Qt.ItemDataRole.EditRole):
            if pd.isna(val):
                return ""
            return str(val)

        style = self._cell_style(r, c)

        if role == QtCore.Qt.ItemDataRole.ForegroundRole and style.get("fg") is not None:
            return QtGui.QBrush(style["fg"])

        if role == QtCore.Qt.ItemDataRole.FontRole and style.get("font") is not None:
            return style["font"]

        if role == QtCore.Qt.ItemDataRole.BackgroundRole:
            qa_bg = style.get("qa_bg")
            if isinstance(qa_bg, QtGui.QColor):
                base_color = qa_bg
            else:
                has_xlsx_bg = bool(style.get("has_xlsx_bg"))
                base_color = style.get("bg")
                if base_color is None:
                    if has_xlsx_bg:
                        base_color = QtGui.QColor(0, 0, 0, 0)
                    else:
                        base_color = column_shade_color(self._is_dark, c)
                        if r % 2 == 1:
                            base_color = base_color.darker(150)

            text = "" if pd.isna(val) else str(val).lower()
            if self.search_term and self.search_term.lower() in text:
                highlight = QtGui.QColor(255, 255, 150, 120)
                if base_color.alpha() == 0:
                    base_color = highlight
                else:
                    base_color = QtGui.QColor(
                        int(base_color.red() * 0.65 + highlight.red() * 0.35),
                        int(base_color.green() * 0.65 + highlight.green() * 0.35),
                        int(base_color.blue() * 0.65 + highlight.blue() * 0.35),
                        base_color.alpha(),
                    )

            return QtGui.QBrush(base_color)
        return None
    
    def headerData(self, section, orientation, role=QtCore.Qt.ItemDataRole.DisplayRole):
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            if orientation == QtCore.Qt.Orientation.Horizontal:
                if section == self.data_column_count():
                    return "+"
                return str(self._df.columns[section])
            else:
                if section == self.data_row_count():
                    return "+"
                return str(self._df.index[section])
        return None

    def flags(self, index):
        if not index.isValid():
            return QtCore.Qt.ItemFlag.NoItemFlags
        if index.column() == self.data_column_count() or index.row() == self.data_row_count():
            return QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled
        return (
            QtCore.Qt.ItemFlag.ItemIsSelectable
            | QtCore.Qt.ItemFlag.ItemIsEnabled
            | QtCore.Qt.ItemFlag.ItemIsEditable
        )

    def setData(self, index, value, role=QtCore.Qt.ItemDataRole.EditRole):
        if role == QtCore.Qt.ItemDataRole.EditRole and index.isValid():
            r, c = index.row(), index.column()
            if r >= self.data_row_count() or c >= self.data_column_count():
                return False
            if isinstance(value, str):
                value = (
                    value.replace("\r\n", "\n")
                    .replace("\n\r", "\n")
                    .replace("\r", "\n")
                    .replace("\u2028", "\n")
                    .replace("\u2029", "\n")
                )
            self._df.iat[r, c] = value
            self._stamp_cell_edit_metadata(r, c)
            self.is_dirty = True
            self.dataChanged.emit(index, index, [QtCore.Qt.ItemDataRole.DisplayRole, QtCore.Qt.ItemDataRole.EditRole])
            self._rebuild_matches()  # update matches after edit
            self.matchesChanged.emit(len(self._matches))
            
            return True
        return False

    # --------------------------------------------------------------------- #
    #  SEARCH & MATCH LOGIC
    # --------------------------------------------------------------------- #
    def set_search_term(self, term: str):
        # Allow spaces, but treat empty or whitespace-only as empty search
        if not term:  # if term is empty string
            term = ""
        # Optional: if you want to treat pure-whitespace as no search:
        # if not term.strip():
        #     term = ""
        if term == self.search_term:
            return
        self.search_term = term
        self._rebuild_matches()
        self.dataChangedHard.emit()
        self.matchesChanged.emit(len(self._matches))

    def _rebuild_matches(self):
        self._matches = []
        if not self.search_term:
            return
        
        if self.case_sensitive:
            st = self.search_term
            for r in range(self.data_row_count()):
                for c in range(self.columnCount() - 1):  # Exclude "+" column
                    val = self._df.iat[r, c]
                    if pd.isna(val):
                        continue
                    if st in str(val):
                        self._matches.append(self.index(r, c))
        else:
            st = self.search_term.lower()
            for r in range(self.data_row_count()):
                for c in range(self.columnCount() - 1):  # Exclude "+" column
                    val = self._df.iat[r, c]
                    if pd.isna(val):
                        continue
                    if st in str(val).lower():
                        self._matches.append(self.index(r, c))

    def match_at(self, pos: int) -> t.Optional[QtCore.QModelIndex]:
        if 0 <= pos < len(self._matches):
            return self._matches[pos]
        return None

    def total_matches(self) -> int:
        return len(self._matches)
    
    def add_new_column(self, col_name: str):
        """Add a new column to the DataFrame"""
        if col_name and col_name not in self._df.columns:
            self.beginInsertColumns(
                QtCore.QModelIndex(),
                self.data_column_count(),
                self.data_column_count(),
            )
            self._df[col_name] = ""
            self.is_dirty = True
            self.endInsertColumns()

    def add_new_row(self):
        """Add a new row to the DataFrame."""
        insert_at = self.data_row_count()
        self.beginInsertRows(QtCore.QModelIndex(), insert_at, insert_at)
        empty_row = pd.DataFrame([[""] * self.data_column_count()], columns=self._df.columns)
        self._df = pd.concat([self._df, empty_row], ignore_index=True)
        self.is_dirty = True
        self.endInsertRows()

    # --------------------------------------------------------------------- #
    #  REPLACE LOGIC (unchanged from your original)
    # --------------------------------------------------------------------- #
    def replace_once(self, start_after: t.Optional[QtCore.QModelIndex]):
        if not self.search_term or self.replace_term is None:
            return None
        matches = self._matches
        if not matches:
            return None
        if start_after and start_after.isValid():
            try:
                idx = matches.index(start_after) + 1
            except ValueError:
                idx = 0
                for i, m in enumerate(matches):
                    if (m.row(), m.column()) > (start_after.row(), start_after.column()):
                        idx = i
                        break
        else:
            idx = 0
        m = matches[idx % len(matches)]
        r, c = m.row(), m.column()
        val = str(self._df.iat[r, c])
        new_val = self._replace_case_insensitive(val, self.search_term, self.replace_term, count=1)
        self._df.iat[r, c] = new_val
        self._stamp_cell_edit_metadata(r, c)
        self.is_dirty = True
        self.dataChanged.emit(m, m, [QtCore.Qt.ItemDataRole.DisplayRole, QtCore.Qt.ItemDataRole.EditRole])
        self._rebuild_matches()
        self.dataChangedHard.emit()
        self.matchesChanged.emit(len(self._matches))
        return m

    def replace_all(self) -> int:
        if not self.search_term:
            return 0
        count = 0
        st = self.search_term
        for r in range(self.data_row_count()):
            for c in range(self.columnCount() - 1):
                val = self._df.iat[r, c]
                if pd.isna(val):
                    continue
                s = str(val)
                if st.lower() in s.lower():
                    new_val = self._replace_case_insensitive(s, st, self.replace_term, count=0)
                    if new_val != s:
                        self._df.iat[r, c] = new_val
                        self._stamp_cell_edit_metadata(r, c)
                        count += 1
        if count:
            self.is_dirty = True
            tl = self.index(0, 0)
            br = self.index(self.rowCount() - 1, self.columnCount() - 1)
            self.dataChanged.emit(tl, br, [QtCore.Qt.ItemDataRole.DisplayRole, QtCore.Qt.ItemDataRole.EditRole])
            self._rebuild_matches()
            self.dataChangedHard.emit()
            self.matchesChanged.emit(len(self._matches))
        return count

    @staticmethod
    def _replace_case_insensitive(text: str, pattern: str, repl: str, count: int) -> str:
        import re
        flags = re.IGNORECASE
        return re.sub(re.escape(pattern), repl, text, count=count if count else 0, flags=flags)


class QAContextTextEdit(QtWidgets.QTextEdit):
    """Editor with a QA-focused context menu for marking selected text errors."""
    qaSpansChanged = QtCore.pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._last_selection: t.Optional[tuple[int, int]] = None
        self._qa_spans: list[dict[str, t.Any]] = []
        self.selectionChanged.connect(self._remember_selection)

    def _remember_selection(self):
        cursor = self.textCursor()
        if not cursor.hasSelection():
            return
        start = int(cursor.selectionStart())
        end = int(cursor.selectionEnd())
        if end > start:
            self._last_selection = (start, end)

    @staticmethod
    def _format_for_kind(kind: str) -> QtGui.QTextCharFormat:
        fmt = QtGui.QTextCharFormat()
        fmt.setBackground(QtGui.QBrush(QA_TEXT_BG_COLOR))
        fmt.setForeground(QtGui.QBrush(QA_TEXT_FG_COLOR))
        if kind == "big":
            fmt.setFontWeight(QtGui.QFont.Weight.Bold)
            fmt.setFontItalic(False)
        else:
            fmt.setFontWeight(QtGui.QFont.Weight.Normal)
            fmt.setFontItalic(True)
        return fmt

    @staticmethod
    def _merge_spans(spans: list[dict[str, t.Any]]) -> list[dict[str, t.Any]]:
        clean: list[dict[str, t.Any]] = []
        for span in spans:
            try:
                start = int(span.get("start", 0))
                length = int(span.get("length", 0))
                kind = str(span.get("kind", ""))
            except Exception:
                continue
            if start < 0 or length <= 0 or kind not in ("big", "small"):
                continue
            clean.append({"start": start, "length": length, "kind": kind})

        if not clean:
            return []
        ordered = sorted(clean, key=lambda s: (int(s["start"]), int(s["length"])))
        merged: list[dict[str, t.Any]] = [ordered[0]]
        for span in ordered[1:]:
            prev = merged[-1]
            prev_end = int(prev["start"]) + int(prev["length"])
            cur_start = int(span["start"])
            cur_end = cur_start + int(span["length"])
            if span["kind"] == prev["kind"] and cur_start <= prev_end:
                prev["length"] = max(prev_end, cur_end) - int(prev["start"])
            else:
                merged.append(span)
        return merged

    def _remove_span_range(self, start: int, end: int):
        updated: list[dict[str, t.Any]] = []
        for span in self._qa_spans:
            s = int(span["start"])
            e = s + int(span["length"])
            if e <= start or s >= end:
                updated.append(dict(span))
                continue
            if s < start:
                updated.append({"start": s, "length": start - s, "kind": span["kind"]})
            if e > end:
                updated.append({"start": end, "length": e - end, "kind": span["kind"]})
        self._qa_spans = self._merge_spans(updated)

    def _replace_span_range(self, start: int, end: int, kind: str):
        self._remove_span_range(start, end)
        self._qa_spans = self._merge_spans(
            self._qa_spans + [{"start": start, "length": end - start, "kind": kind}]
        )

    def contextMenuEvent(self, event: QtGui.QContextMenuEvent):
        cursor = self.textCursor()
        has_selection = bool(cursor.hasSelection())
        if has_selection:
            sel_start = int(cursor.selectionStart())
            sel_end = int(cursor.selectionEnd())
        elif self._last_selection is not None:
            sel_start, sel_end = self._last_selection
            has_selection = sel_end > sel_start
        else:
            sel_start = -1
            sel_end = -1

        menu = QtWidgets.QMenu(self)

        mark_big = menu.addAction("◉ Mark Big Issue")
        mark_small = menu.addAction("● Mark Small Issue")
        clear_issue = menu.addAction("Clear Issue")
        mark_big.setEnabled(has_selection)
        mark_small.setEnabled(has_selection)
        clear_issue.setEnabled(has_selection)

        chosen = menu.exec(event.globalPos())
        if chosen == mark_big:
            self._apply_error_mark(sel_start, sel_end, big=True)
        elif chosen == mark_small:
            self._apply_error_mark(sel_start, sel_end, big=False)
        elif chosen == clear_issue:
            self._clear_error_mark(sel_start, sel_end)

    def _apply_error_mark(self, start: int, end: int, big: bool):
        if start < 0 or end <= start:
            return

        self._last_selection = (start, end)
        kind = "big" if big else "small"
        self._replace_span_range(start, end, kind)
        cursor = QtGui.QTextCursor(self.document())
        cursor.setPosition(start)
        cursor.setPosition(end, QtGui.QTextCursor.MoveMode.KeepAnchor)

        fmt = self._format_for_kind(kind)
        cursor.mergeCharFormat(fmt)
        self.qaSpansChanged.emit()

    def _clear_error_mark(self, start: int, end: int):
        if start < 0 or end <= start:
            return

        self._last_selection = (start, end)
        self._remove_span_range(start, end)

        cursor = QtGui.QTextCursor(self.document())
        cursor.setPosition(start)
        cursor.setPosition(end, QtGui.QTextCursor.MoveMode.KeepAnchor)

        clear_fmt = QtGui.QTextCharFormat()
        clear_fmt.setFontWeight(QtGui.QFont.Weight.Normal)
        clear_fmt.setFontItalic(False)
        if hasattr(clear_fmt, "clearBackground"):
            clear_fmt.clearBackground()
        else:
            clear_fmt.setBackground(QtGui.QBrush())
        if hasattr(clear_fmt, "clearForeground"):
            clear_fmt.clearForeground()
        else:
            clear_fmt.setForeground(QtGui.QBrush())
        cursor.mergeCharFormat(clear_fmt)
        self.qaSpansChanged.emit()

    def extract_qa_spans(self) -> list[dict[str, t.Any]]:
        return [dict(span) for span in self._merge_spans(self._qa_spans)]

    def apply_qa_spans(self, spans: list[dict[str, t.Any]]):
        self._qa_spans = self._merge_spans(spans or [])
        if not self._qa_spans:
            return
        doc_len = self.document().characterCount()
        for span in self._qa_spans:
            try:
                start = max(0, int(span.get("start", 0)))
                length = max(0, int(span.get("length", 0)))
                kind = str(span.get("kind", ""))
            except Exception:
                continue
            if length <= 0 or kind not in ("big", "small"):
                continue
            end = min(doc_len - 1, start + length)
            if end <= start:
                continue

            cursor = self.textCursor()
            cursor.setPosition(start)
            cursor.setPosition(end, QtGui.QTextCursor.MoveMode.KeepAnchor)
            cursor.mergeCharFormat(self._format_for_kind(kind))


class HighlightDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, model: DataFrameModel):
        super().__init__()
        self.model = model
        
        # Cache font metrics
        self._metrics_cache = {}

    def _persist_editor_qa_spans(self, editor: QAContextTextEdit, index):
        if not isinstance(self.model, DataFrameModel):
            return
        if not index.isValid():
            return
        self.model.set_cell_qa_spans(
            index.row(),
            index.column(),
            editor.extract_qa_spans(),
            emit_data_changed=False,
        )
    
    def sizeHint(self, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex) -> QtCore.QSize:
        text = index.data(QtCore.Qt.ItemDataRole.DisplayRole) or ""
        cell_font = index.data(QtCore.Qt.ItemDataRole.FontRole) or option.font
        
        # Fast path: short text without newlines
        if len(text) < 100 and '\n' not in text:
            metrics = QtGui.QFontMetrics(cell_font)
            width = metrics.horizontalAdvance(text) + 28  # padding
            height = metrics.lineSpacing() + TABLE_VERTICAL_TEXT_PADDING
            return QtCore.QSize(width, height)
        
        # Slow path for multiline/long text
        doc = QtGui.QTextDocument()
        doc.setDefaultFont(cell_font)
        doc.setPlainText(text)
        doc.setTextWidth(option.rect.width())
        _apply_document_line_spacing(doc)
        
        content_height = doc.size().height()
        total_height = int(content_height + TABLE_VERTICAL_TEXT_PADDING)
        
        return QtCore.QSize(int(doc.idealWidth()), total_height)


    def paint(self, painter: QtGui.QPainter, option: QtWidgets.QStyleOptionViewItem, index: QtCore.QModelIndex):
        option = QtWidgets.QStyleOptionViewItem(option)
        self.initStyleOption(option, index)
        painter.save()
        
        # Get base background color
        bg_brush = index.data(QtCore.Qt.ItemDataRole.BackgroundRole)
        fg_brush = index.data(QtCore.Qt.ItemDataRole.ForegroundRole)
        cell_font = index.data(QtCore.Qt.ItemDataRole.FontRole) or option.font
        
        # Check if item is selected
        is_selected = option.state & QtWidgets.QStyle.StateFlag.State_Selected
        
        if bg_brush:
            base_color = bg_brush.color()
            if is_selected:
                # Make background lighter when selected (add 40 to each RGB component)
                lighter_color = QtGui.QColor(
                    min(255, base_color.red() - 20),
                    min(255, base_color.green() - 20),
                    min(255, base_color.blue() - 20),
                    base_color.alpha()
                )
                painter.fillRect(option.rect, lighter_color)
            else:
                painter.fillRect(option.rect, bg_brush)
        elif is_selected:
            # Fallback to standard selection color if no background
            painter.fillRect(option.rect, option.palette.highlight())
        
        text = index.data(QtCore.Qt.ItemDataRole.DisplayRole) or ""
        st = self.model.search_term
        if st:
            if self.model.case_sensitive:
                has_term = st in text
            else:
                has_term = st.lower() in text.lower()
        else:
            has_term = False
        
        rect = option.rect.adjusted(4, 2, -4, -2)
        
        # Create document for text (rest of your existing logic)
        doc = QtGui.QTextDocument()
        doc.setDefaultFont(cell_font)
        text_option = QtGui.QTextOption()
        text_option.setWrapMode(QtGui.QTextOption.WrapMode.WrapAtWordBoundaryOrAnywhere)
        text_option.setAlignment(QtCore.Qt.AlignmentFlag.AlignVCenter | QtCore.Qt.AlignmentFlag.AlignLeft)
        doc.setDefaultTextOption(text_option)
        doc.setPlainText(text)
        doc.setTextWidth(rect.width())
        _apply_document_line_spacing(doc)

        base_text_format = QtGui.QTextCharFormat()
        if is_selected:
            base_text_format.setForeground(option.palette.highlightedText())
        elif fg_brush:
            fg_color = _invert_black_white_display_color(fg_brush.color())
            base_text_format.setForeground(QtGui.QBrush(fg_color))
        else:
            if self.model._is_dark:
                base_text_format.setForeground(QtGui.QBrush(QtGui.QColor(235, 235, 235)))
            else:
                base_text_format.setForeground(option.palette.text())

        base_cursor = QtGui.QTextCursor(doc)
        base_cursor.select(QtGui.QTextCursor.SelectionType.Document)
        base_cursor.mergeCharFormat(base_text_format)

        qa_spans = self.model.cell_qa_spans(index.row(), index.column())
        if qa_spans:
            doc_len = doc.characterCount()
            for span in qa_spans:
                try:
                    start = max(0, int(span.get("start", 0)))
                    length = max(0, int(span.get("length", 0)))
                    kind = str(span.get("kind", ""))
                except Exception:
                    continue
                if length <= 0 or kind not in ("big", "small"):
                    continue
                end = min(doc_len - 1, start + length)
                if end <= start:
                    continue

                qa_fmt = QtGui.QTextCharFormat()
                qa_fmt.setBackground(QtGui.QBrush(QA_TEXT_BG_COLOR))
                qa_fmt.setForeground(QtGui.QBrush(QA_TEXT_FG_COLOR))
                if kind == "big":
                    qa_fmt.setFontWeight(QtGui.QFont.Weight.Bold)
                    qa_fmt.setFontItalic(False)
                else:
                    qa_fmt.setFontWeight(QtGui.QFont.Weight.Normal)
                    qa_fmt.setFontItalic(True)

                qa_cursor = QtGui.QTextCursor(doc)
                qa_cursor.setPosition(start)
                qa_cursor.setPosition(end, QtGui.QTextCursor.MoveMode.KeepAnchor)
                qa_cursor.mergeCharFormat(qa_fmt)
        
        # Highlight matching words (existing logic unchanged)
        if has_term:
            start = 0
            highlights = []
            
            if self.model.case_sensitive:
                search_text = st
                search_in = text
            else:
                search_text = st.lower()
                search_in = text.lower()
            
            text_len = len(st)
            
            while True:
                i = search_in.find(search_text, start)
                if i == -1:
                    break
                highlights.append((i, text_len))
                start = i + text_len
            
            if highlights:
                highlight_format = QtGui.QTextCharFormat()
                highlight_format.setBackground(QtGui.QBrush(QtGui.QColor(0, 200, 0, 140)))
                highlight_format.setForeground(QtGui.QBrush(QtCore.Qt.GlobalColor.white)) # Optional: ensure text is readable

                # APPLY FORMAT TO EACH MATCH
                for pos, length in highlights:
                    cursor = QtGui.QTextCursor(doc)
                    cursor.setPosition(pos)
                    cursor.setPosition(pos + length, QtGui.QTextCursor.MoveMode.KeepAnchor)
                    cursor.mergeCharFormat(highlight_format)

        painter.translate(rect.topLeft())
        clip = QtCore.QRectF(0, 0, rect.width(), rect.height())
        doc.drawContents(painter, clip)
        
        painter.restore()

        # Draw editor marker (small colored dot) in the lower-right corner.
        if (
            index.row() < self.model.data_row_count()
            and index.column() < self.model.data_column_count()
        ):
            edit_meta = self.model.cell_edit_metadata(index.row(), index.column())
            if edit_meta is not None:
                edited_by, _ = edit_meta
                marker_color = _edited_cell_color_for_username(edited_by)
                if option.rect.width() >= 10 and option.rect.height() >= 10:
                    dot_radius = 3.6
                    center = QtCore.QPointF(
                        float(option.rect.right() - 5),
                        float(option.rect.bottom() - 5),
                    )
                    painter.save()
                    painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
                    painter.setPen(QtGui.QPen(marker_color.darker(175), 1))
                    painter.setBrush(QtGui.QBrush(marker_color))
                    painter.drawEllipse(center, dot_radius, dot_radius)
                    painter.restore()
        
        
    def createEditor(self, parent, option, index):
        editor = QAContextTextEdit(parent)
        editor.setAcceptRichText(False)
        cell_font = index.data(QtCore.Qt.ItemDataRole.FontRole) or option.font
        editor.setFont(cell_font)
        editor.setLineWrapMode(QtWidgets.QTextEdit.LineWrapMode.WidgetWidth)
        editor.setWordWrapMode(QtGui.QTextOption.WrapMode.WrapAtWordBoundaryOrAnywhere)
        editor.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
        editor.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        editor.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        editor.document().setDocumentMargin(1.0)
        _apply_document_line_spacing(editor.document())
        editor.setViewportMargins(4, 2, 4, 2)
        editor.setStyleSheet("QTextEdit { border: none; padding: 0px; }")
        persistent_index = QtCore.QPersistentModelIndex(index)
        editor.qaSpansChanged.connect(
            lambda ed=editor, idx=persistent_index: self._persist_editor_qa_spans(ed, idx)
        )
        return editor

    def setEditorData(self, editor, index):
        cell_font = index.data(QtCore.Qt.ItemDataRole.FontRole)
        if cell_font:
            editor.setFont(cell_font)
        editor.setPlainText(index.data(QtCore.Qt.ItemDataRole.DisplayRole) or "")
        _apply_document_line_spacing(editor.document())
        if isinstance(editor, QAContextTextEdit) and isinstance(self.model, DataFrameModel):
            editor.apply_qa_spans(self.model.cell_qa_spans(index.row(), index.column()))
        cursor = editor.textCursor()
        cursor.movePosition(QtGui.QTextCursor.MoveOperation.Start)
        editor.setTextCursor(cursor)

    def updateEditorGeometry(self, editor, option, index):
        # Use full cell geometry to avoid extra narrowing/wrapping in edit mode.
        editor.setGeometry(option.rect)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QAContextTextEdit) and isinstance(model, DataFrameModel):
            model.set_cell_qa_spans(index.row(), index.column(), editor.extract_qa_spans())
        model.setData(index, editor.toPlainText(), QtCore.Qt.ItemDataRole.EditRole)
class DragDropWidget(QtWidgets.QFrame):
    fileDropped = QtCore.pyqtSignal(Path)

    def mousePressEvent(self, e: QtGui.QMouseEvent):
        if getattr(self, "_drop_enabled", True) and e.button() == QtCore.Qt.MouseButton.LeftButton:
            if self.icon.geometry().contains(e.position().toPoint()):
                parent = self.parentWidget()
                while parent is not None and not hasattr(parent, "open_file_dialog"):
                    parent = parent.parentWidget()
                if parent is not None:
                    parent.open_file_dialog()
        super().mousePressEvent(e)

    def __init__(self):
        super().__init__()
        self._drop_enabled = True
        self.setAcceptDrops(True)
        self.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.icon = QtWidgets.QLabel(DROP_HINT_TEXT)
        self.icon.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.icon.setStyleSheet("font-size: 16px;")
        layout.addWidget(self.icon)


    def set_drop_enabled(self, enabled: bool):
        self._drop_enabled = bool(enabled)
        self.setAcceptDrops(self._drop_enabled)

    def dragEnterEvent(self, event: QtGui.QDragEnterEvent):
        if not self._drop_enabled:
            event.ignore()
            return
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    suffix = Path(url.toLocalFile()).suffix.lower()
                    if suffix in [".csv", ".xls", ".xlsx", ".jsonl"]:
                        event.acceptProposedAction()
                        return
        event.ignore()

    def dropEvent(self, event: QtGui.QDropEvent):
        if not self._drop_enabled:
            event.ignore()
            return
        for url in event.mimeData().urls():
            if url.isLocalFile():
                path = Path(url.toLocalFile())
                if path.suffix.lower() in [".csv", ".xls", ".xlsx", ".jsonl"]:
                    self.fileDropped.emit(path)
                    return


class WindowControls(QtWidgets.QWidget):
    minimizeRequested = QtCore.pyqtSignal()
    maximizeRestoreRequested = QtCore.pyqtSignal()
    closeRequested = QtCore.pyqtSignal()
    backRequested = QtCore.pyqtSignal()
    shareRequested = QtCore.pyqtSignal()

    def __init__(self, title: str = ""):
        super().__init__()
        self._drag_start_pos = None
        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        
        self.back_btn = QtWidgets.QToolButton()
        self.back_btn.setText("← Back")
        self.back_btn.clicked.connect(self.backRequested.emit)

        self.title_lbl = QtWidgets.QLabel(title)
        self.title_lbl.setStyleSheet("font-weight: 600;")
        self.title_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        self.share_btn = QtWidgets.QToolButton()
        self.share_btn.setText("\u2197 Share")
        self.share_btn.setToolTip("Share")
        self.share_btn.clicked.connect(self.shareRequested.emit)

        self.share_count_lbl = QtWidgets.QLabel("0")
        self.share_count_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.share_count_lbl.setMinimumWidth(20)
        self.share_count_lbl.setStyleSheet(
            "font-weight: 600; color: rgb(190, 190, 190); padding: 0px 4px;"
        )

        self.min_btn = QtWidgets.QToolButton()
        self.min_btn.setText("—")
        self.min_btn.setToolTip("Minimize")
        self.min_btn.clicked.connect(self.minimizeRequested.emit)

        self.max_btn = QtWidgets.QToolButton()
        self.max_btn.setText("▢")
        self.max_btn.setToolTip("Maximize/Restore")
        self.max_btn.clicked.connect(self.maximizeRestoreRequested.emit)

        self.close_btn = QtWidgets.QToolButton()
        self.close_btn.setText("✕")
        self.close_btn.setToolTip("Close")
        self.close_btn.clicked.connect(self.closeRequested.emit)

        self._left_balance = QtWidgets.QWidget()
        self._left_balance.setFixedWidth(0)
        self._left_balance.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Fixed,
            QtWidgets.QSizePolicy.Policy.Preferred,
        )
        self._right_balance = QtWidgets.QWidget()
        self._right_balance.setFixedWidth(0)
        self._right_balance.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Fixed,
            QtWidgets.QSizePolicy.Policy.Preferred,
        )

        # Left side: back button
        layout.addWidget(self.back_btn)
        layout.addWidget(self._left_balance)
        layout.addStretch(1)
        
        # Center: title
        layout.addWidget(self.title_lbl)
        layout.addStretch(1)
        layout.addWidget(self._right_balance)
        
        # Right side: share + live count + window controls
        layout.addWidget(self.share_btn)
        layout.addWidget(self.share_count_lbl)
        for b in (self.min_btn, self.max_btn, self.close_btn):
            layout.addWidget(b)

    @staticmethod
    def _visible_widget_width(widget: QtWidgets.QWidget) -> int:
        if widget is None or not widget.isVisible():
            return 0
        return int(widget.sizeHint().width())

    def sync_title_balance(self):
        left_width = self._visible_widget_width(self.back_btn)
        right_width = (
            self._visible_widget_width(self.share_btn)
            + self._visible_widget_width(self.share_count_lbl)
            + self._visible_widget_width(self.min_btn)
            + self._visible_widget_width(self.max_btn)
            + self._visible_widget_width(self.close_btn)
        )
        self._left_balance.setFixedWidth(max(0, right_width - left_width))
        self._right_balance.setFixedWidth(max(0, left_width - right_width))

    def resizeEvent(self, event: QtGui.QResizeEvent):
        super().resizeEvent(event)
        self.sync_title_balance()
            
    def mousePressEvent(self, event: QtGui.QMouseEvent):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self._drag_start_pos = event.globalPosition().toPoint()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QtGui.QMouseEvent):
        if self._drag_start_pos and event.buttons() & QtCore.Qt.MouseButton.LeftButton:
            delta = event.globalPosition().toPoint() - self._drag_start_pos
            self.window().move(self.window().pos() + delta)
            self._drag_start_pos = event.globalPosition().toPoint()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QtGui.QMouseEvent):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self._drag_start_pos = None
        super().mouseReleaseEvent(event)
        
class AnimatedScrollBar(QtWidgets.QScrollBar):
    """A scrollbar that animates value changes for buttery-smooth scrolling"""
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self._animation = QtCore.QPropertyAnimation(self, b"value")
        self._animation.setDuration(300)  # 180ms for smooth but responsive feel
        self._animation.setEasingCurve(QtCore.QEasingCurve.Type.OutQuad)  # Natural deceleration
        self._is_dragging = False
        
    def setValue(self, value):
        # If user is dragging the handle, update instantly for immediate feedback
        if self._is_dragging:
            super().setValue(value)
            return
            
        # If same value, nothing to do
        if value == self.value():
            return
            
        # Stop any ongoing animation before starting new one
        if self._animation.state() == QtCore.QPropertyAnimation.State.Running:
            self._animation.stop()
            
        # Animate from current to target value
        self._animation.setStartValue(self.value())
        self._animation.setEndValue(value)
        self._animation.start()
        
    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self._is_dragging = True
            self._animation.stop()  # Stop animation when user grabs handle
        super().mousePressEvent(event)
        
    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self._is_dragging = False
        super().mouseReleaseEvent(event)

    def setValueInstant(self, value):
            """Force set value immediately, cancelling any animation."""
            if self._animation.state() == QtCore.QPropertyAnimation.State.Running:
                self._animation.stop()
            super().setValue(value)


class SmoothScrollTableView(QtWidgets.QTableView):
    """QTableView with smooth animated scrolling for both wheel and scrollbar interactions
    
    Features:
    - Animated scrolling via AnimatedScrollBar
    - Wheel event smoothing
    - 2D Middle-click autoscroll (NEW) - hold middle button and move mouse to scroll in any direction
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        # Enable per-pixel scrolling
        self.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollMode.ScrollPerPixel)
        
        # Replace default scrollbars with animated versions
        v_scrollbar = AnimatedScrollBar(QtCore.Qt.Orientation.Vertical, self)
        h_scrollbar = AnimatedScrollBar(QtCore.Qt.Orientation.Horizontal, self)
        
        self.setVerticalScrollBar(v_scrollbar)
        self.setHorizontalScrollBar(h_scrollbar)
        
        # --- 2D Middle-click autoscroll support ---
        self._autoscroll_active = False
        self._autoscroll_origin = None
        self._autoscroll_speed_x = 0
        self._autoscroll_speed_y = 0
        self._autoscroll_deadzone = 15  # pixels before scrolling starts
        self._autoscroll_sensitivity = 0.35  # INCREASED for faster scrolling
        
        # Timer for smooth autoscroll updates
        self._autoscroll_timer = QtCore.QTimer()
        self._autoscroll_timer.setInterval(16)  # ~60fps
        self._autoscroll_timer.timeout.connect(self._update_autoscroll)
        self._remote_selections: dict[str, tuple[int, int]] = {}

    def set_remote_selections(self, selections: dict[str, tuple[int, int]]):
        normalized: dict[str, tuple[int, int]] = {}
        for raw_name, raw_pos in (selections or {}).items():
            name = _normalize_username(raw_name)
            if not name:
                continue
            if not isinstance(raw_pos, (tuple, list)) or len(raw_pos) != 2:
                continue
            try:
                row = int(raw_pos[0])
                col = int(raw_pos[1])
            except Exception:
                continue
            normalized[name] = (row, col)

        if normalized != self._remote_selections:
            self._remote_selections = normalized
            self.viewport().update()

    def clear_remote_selections(self):
        if self._remote_selections:
            self._remote_selections.clear()
            self.viewport().update()
        
    def mousePressEvent(self, event: QtGui.QMouseEvent):
        # Any button press while autoscrolling should stop it
        if self._autoscroll_active:
            self._stop_autoscroll()
            
        # Middle button starts 2D autoscroll
        if event.button() == QtCore.Qt.MouseButton.MiddleButton:
            v_scrollbar = self.verticalScrollBar()
            h_scrollbar = self.horizontalScrollBar()
            
            # Only enable if there's actual content to scroll (vertical OR horizontal)
            has_vertical = v_scrollbar.maximum() > v_scrollbar.minimum()
            has_horizontal = h_scrollbar.maximum() > h_scrollbar.minimum()
            
            if has_vertical or has_horizontal:
                self._autoscroll_active = True
                self._autoscroll_origin = event.position().toPoint()
                self.viewport().setCursor(QtCore.Qt.CursorShape.SizeAllCursor)  # 2D cursor
                self._autoscroll_timer.start()
                event.accept()
                return
            
        super().mousePressEvent(event)
        
    def mouseReleaseEvent(self, event: QtGui.QMouseEvent):
        # Middle button stops autoscroll
        if event.button() == QtCore.Qt.MouseButton.MiddleButton and self._autoscroll_active:
            self._stop_autoscroll()
            event.accept()
            return
            
        super().mouseReleaseEvent(event)
        
    def mouseMoveEvent(self, event: QtGui.QMouseEvent):
        # Update autoscroll speed based on mouse position (2D)
        if self._autoscroll_active and self._autoscroll_origin:
            # Calculate distance from origin for both axes
            delta_x = event.position().x() - self._autoscroll_origin.x()
            delta_y = event.position().y() - self._autoscroll_origin.y()
            
            # Apply deadzone to horizontal
            if abs(delta_x) < self._autoscroll_deadzone:
                self._autoscroll_speed_x = 0
            else:
                effective_delta_x = delta_x - (self._autoscroll_deadzone * (1 if delta_x > 0 else -1))
                self._autoscroll_speed_x = effective_delta_x * self._autoscroll_sensitivity
            
            # Apply deadzone to vertical
            if abs(delta_y) < self._autoscroll_deadzone:
                self._autoscroll_speed_y = 0
            else:
                effective_delta_y = delta_y - (self._autoscroll_deadzone * (1 if delta_y > 0 else -1))
                self._autoscroll_speed_y = effective_delta_y * self._autoscroll_sensitivity
            
        super().mouseMoveEvent(event)
        
    def leaveEvent(self, event: QtCore.QEvent):
        # Stop autoscroll if mouse leaves viewport
        if self._autoscroll_active:
            self._stop_autoscroll()
        super().leaveEvent(event)
        
    def _update_autoscroll(self):
        """Update scroll position based on autoscroll speed (both axes)"""
        if not self._autoscroll_active:
            return
            
        # Update vertical scrollbar - bypass animation
        if self._autoscroll_speed_y != 0:
            v_scrollbar = self.verticalScrollBar()
            current_value_y = v_scrollbar.value()
            new_value_y = current_value_y + self._autoscroll_speed_y
            new_value_y = max(v_scrollbar.minimum(), min(new_value_y, v_scrollbar.maximum()))
            
            # Stop any ongoing animation and set directly
            if hasattr(v_scrollbar, '_animation'):
                v_scrollbar._animation.stop()
            QtWidgets.QScrollBar.setValue(v_scrollbar, int(new_value_y))  # Call parent class method
        
        # Update horizontal scrollbar - bypass animation
        if self._autoscroll_speed_x != 0:
            h_scrollbar = self.horizontalScrollBar()
            current_value_x = h_scrollbar.value()
            new_value_x = current_value_x + self._autoscroll_speed_x
            new_value_x = max(h_scrollbar.minimum(), min(new_value_x, h_scrollbar.maximum()))
            
            # Stop any ongoing animation and set directly
            if hasattr(h_scrollbar, '_animation'):
                h_scrollbar._animation.stop()
            QtWidgets.QScrollBar.setValue(h_scrollbar, int(new_value_x))  # Call parent class method
            
    def _stop_autoscroll(self):
        """Stop autoscroll and reset state"""
        if not self._autoscroll_active:
            return
            
        self._autoscroll_active = False
        self._autoscroll_origin = None
        self._autoscroll_speed_x = 0
        self._autoscroll_speed_y = 0
        self._autoscroll_timer.stop()
        self.viewport().unsetCursor()
        self.viewport().update()  # Clear any marker

    def _paint_remote_selections(self):
        if not self._remote_selections:
            return

        model = self.model()
        if model is None:
            return

        max_row = model.rowCount() - 1
        max_col = model.columnCount() - 1
        if max_row <= 0 or max_col <= 0:
            return

        grouped: dict[tuple[int, int], list[str]] = {}
        for username, (row, col) in self._remote_selections.items():
            if row < 0 or col < 0 or row >= max_row or col >= max_col:
                continue
            grouped.setdefault((row, col), []).append(username)

        if not grouped:
            return

        painter = QtGui.QPainter(self.viewport())
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)

        label_font = QtGui.QFont(self.font())
        label_font.setBold(True)
        label_font.setPointSize(max(7, label_font.pointSize() - 2))
        painter.setFont(label_font)
        metrics = QtGui.QFontMetrics(label_font)

        for (row, col), users in grouped.items():
            idx = model.index(row, col)
            rect = self.visualRect(idx)
            if not rect.isValid() or rect.isEmpty():
                continue
            if not self.viewport().rect().intersects(rect):
                continue

            users_sorted = sorted(users, key=lambda s: s.lower())
            anchor_user = users_sorted[0]
            color = _presence_color_for_username(anchor_user)

            fill_color = QtGui.QColor(color)
            fill_color.setAlpha(54)
            border_color = QtGui.QColor(color)
            border_color.setAlpha(220)

            inner = rect.adjusted(1, 1, -1, -1)
            painter.fillRect(inner, fill_color)
            painter.setPen(QtGui.QPen(border_color, 2))
            painter.setBrush(QtCore.Qt.BrushStyle.NoBrush)
            painter.drawRect(inner)

            if len(users_sorted) == 1:
                label_text = users_sorted[0]
            else:
                label_text = f"{users_sorted[0]} +{len(users_sorted) - 1}"

            if rect.width() < 26 or rect.height() < 14:
                continue

            badge_height = max(13, metrics.height() + 2)
            max_badge_width = max(24, rect.width() - 4)
            text = metrics.elidedText(
                label_text,
                QtCore.Qt.TextElideMode.ElideRight,
                max_badge_width - 8,
            )
            badge_width = min(max_badge_width, metrics.horizontalAdvance(text) + 8)
            badge_x = rect.right() - badge_width - 2
            badge_y = rect.top() + 2
            if badge_y + badge_height > rect.bottom():
                badge_y = rect.bottom() - badge_height - 1

            badge_rect = QtCore.QRect(badge_x, badge_y, badge_width, badge_height)
            badge_fill = QtGui.QColor(color)
            badge_fill.setAlpha(235)
            painter.setPen(QtGui.QPen(badge_fill.darker(150), 1))
            painter.setBrush(badge_fill)
            painter.drawRoundedRect(badge_rect, 3, 3)

            painter.setPen(QtGui.QPen(QtCore.Qt.GlobalColor.white))
            painter.drawText(
                badge_rect.adjusted(4, 0, -4, 0),
                QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter,
                text,
            )

    def paintEvent(self, event: QtGui.QPaintEvent):
        """Custom paint to draw 2D autoscroll marker when active"""
        super().paintEvent(event)
        self._paint_remote_selections()
        
        if self._autoscroll_active and self._autoscroll_origin:
            painter = QtGui.QPainter(self.viewport())
            painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
            
            marker_size = 20
            rect = QtCore.QRect(
                self._autoscroll_origin.x() - marker_size // 2,
                self._autoscroll_origin.y() - marker_size // 2,
                marker_size,
                marker_size
            )
            
            # Draw crosshair circle
            painter.setPen(QtGui.QPen(QtGui.QColor(100, 150, 255, 180), 2))
            painter.setBrush(QtGui.QColor(100, 150, 255, 60))
            painter.drawEllipse(rect)
            
            # Draw direction arrows with higher visibility
            arrow_size = 7  # Slightly larger
            painter.setBrush(QtGui.QColor(100, 200, 255, 220))  # Brighter
            
            # Vertical arrow
            if abs(self._autoscroll_speed_y) > 2:
                if self._autoscroll_speed_y < 0:  # Up arrow
                    points = [
                        QtCore.QPoint(self._autoscroll_origin.x(), self._autoscroll_origin.y() - marker_size // 2 - 3),
                        QtCore.QPoint(self._autoscroll_origin.x() - arrow_size, self._autoscroll_origin.y() - marker_size // 2 + 4),
                        QtCore.QPoint(self._autoscroll_origin.x() + arrow_size, self._autoscroll_origin.y() - marker_size // 2 + 4)
                    ]
                else:  # Down arrow
                    points = [
                        QtCore.QPoint(self._autoscroll_origin.x(), self._autoscroll_origin.y() + marker_size // 2 + 3),
                        QtCore.QPoint(self._autoscroll_origin.x() - arrow_size, self._autoscroll_origin.y() + marker_size // 2 - 4),
                        QtCore.QPoint(self._autoscroll_origin.x() + arrow_size, self._autoscroll_origin.y() + marker_size // 2 - 4)
                    ]
                painter.drawPolygon(QtGui.QPolygon(points))
            
            # Horizontal arrow
            if abs(self._autoscroll_speed_x) > 2:
                if self._autoscroll_speed_x < 0:  # Left arrow
                    points = [
                        QtCore.QPoint(self._autoscroll_origin.x() - marker_size // 2 - 3, self._autoscroll_origin.y()),
                        QtCore.QPoint(self._autoscroll_origin.x() - marker_size // 2 + 4, self._autoscroll_origin.y() - arrow_size),
                        QtCore.QPoint(self._autoscroll_origin.x() - marker_size // 2 + 4, self._autoscroll_origin.y() + arrow_size)
                    ]
                else:  # Right arrow
                    points = [
                        QtCore.QPoint(self._autoscroll_origin.x() + marker_size // 2 + 3, self._autoscroll_origin.y()),
                        QtCore.QPoint(self._autoscroll_origin.x() + marker_size // 2 - 4, self._autoscroll_origin.y() - arrow_size),
                        QtCore.QPoint(self._autoscroll_origin.x() + marker_size // 2 - 4, self._autoscroll_origin.y() + arrow_size)
                    ]
                painter.drawPolygon(QtGui.QPolygon(points))
        
    def wheelEvent(self, event):
        # Animate mouse wheel scrolling for vertical movement
        if event.modifiers() == QtCore.Qt.KeyboardModifier.NoModifier:
            delta = event.angleDelta().y()
            scrollbar = self.verticalScrollBar()
            
            # Calculate scroll distance (3 lines per wheel step for reasonable speed)
            single_step = scrollbar.singleStep()
            if delta > 0:
                target_value = max(scrollbar.minimum(), scrollbar.value() - single_step * 0.75)
            else:
                target_value = min(scrollbar.maximum(), scrollbar.value() + single_step * 0.75)
                
            # Animate to target value (this triggers AnimatedScrollBar.setValue)
            scrollbar.setValue(target_value)
            event.accept()
        else:
            # Handle horizontal scrolling or Ctrl+wheel (zoom) normally
            super().wheelEvent(event)


class DataViewerPage(QtWidgets.QWidget):
    backRequested = QtCore.pyqtSignal()

    def __init__(self, app_ref: 'MainWindow'):
        super().__init__()
        self.setAcceptDrops(True)
        self.app_ref = app_ref
        self.current_path: t.Optional[Path] = None
        self.model: t.Optional[DataFrameModel] = None
        self.current_match_pos: int = -1
        self._xlsx_sheet_names: list[str] = []
        self._sheet_reload_in_progress = False
        self._session_role = "local"
        self._local_username = _read_local_os_username()
        self._collab_server: t.Optional[CollaborationServer] = None
        self._collab_client: t.Optional[CollaborationClient] = None
        self._collab_session_token = ""
        self._collab_share_link = ""
        self._collab_client_connected = False
        self._collab_connecting = False
        self._collab_connected_guest_count = 0
        self._collab_remote_selections: dict[str, tuple[int, int]] = {}
        self._collab_last_sent_selection: t.Optional[tuple[int, int]] = None
        self._table_selection_model: t.Optional[QtCore.QItemSelectionModel] = None
        self._collab_last_notice_ts = 0.0
        self._collab_is_applying_remote_snapshot = False
        self._collab_revision = 0
        self._collab_broadcast_timer = QtCore.QTimer(self)
        self._collab_broadcast_timer.setSingleShot(True)
        self._collab_broadcast_timer.setInterval(COLLAB_BROADCAST_DEBOUNCE_MS)
        self._collab_broadcast_timer.timeout.connect(self._broadcast_current_snapshot)

        outer = QtWidgets.QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Window controls bar
        self.controls = WindowControls(APP_NAME)
        self.controls.backRequested.connect(self.backRequested.emit)
        self.controls.back_btn.setAutoRaise(False)
        self.controls.back_btn.setStyleSheet(
            """
            QToolButton:disabled {
                color: rgb(120, 120, 120);
                background-color: rgb(55, 55, 55);
                border: 1px solid rgb(75, 75, 75);
                border-radius: 4px;
                padding: 2px 8px;
            }
            """
        )
        self.controls.back_btn.setEnabled(False)
        self.controls.minimizeRequested.connect(lambda: self.window().showMinimized())
        self.controls.maximizeRestoreRequested.connect(self._toggle_max_restore)
        self.controls.closeRequested.connect(QtWidgets.QApplication.instance().quit)
        self.controls.share_btn.setToolTip("Start hosting and copy share link")
        self.controls.share_btn.setVisible(False)
        self.controls.share_btn.setEnabled(False)
        self.controls.share_count_lbl.setVisible(False)
        self.controls.shareRequested.connect(self._on_share_requested)
        self.controls.setFixedHeight(40)

        # Toolbar
        toolbar = QtWidgets.QToolBar()
        toolbar.setIconSize(QtCore.QSize(16, 16))
        toolbar.setMovable(False)
        toolbar.setFloatable(False)
        self.toolbar = toolbar

        # --- Actions ---
        self.btn_open = QtGui.QAction("Open", self)
        self.btn_open.triggered.connect(self.open_file_dialog)
        toolbar.addAction(self.btn_open)
        toolbar.addSeparator()

        self.btn_save = QtGui.QAction("Save", self)
        self.btn_save.triggered.connect(self.save_file)
        toolbar.addAction(self.btn_save)
        toolbar.addSeparator()

        self.btn_save_copy = QtGui.QAction("Save Copy", self)
        self.btn_save_copy.triggered.connect(self.save_copy)
        toolbar.addAction(self.btn_save_copy)

        toolbar.addSeparator()

        self.sheet_label = QtWidgets.QLabel("Sheet:")
        self.sheet_label.setVisible(True)
        toolbar.addWidget(self.sheet_label)

        self.sheet_combo = QtWidgets.QComboBox()
        self.sheet_combo.setMinimumWidth(120)
        self.sheet_combo.setMaximumWidth(240)
        self.sheet_combo.setSizeAdjustPolicy(QtWidgets.QComboBox.SizeAdjustPolicy.AdjustToContents)
        self.sheet_combo.setVisible(True)
        self.sheet_combo.addItem("(No sheet)")
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        toolbar.addWidget(self.sheet_combo)

        toolbar.addSeparator()

        # --- Dark/Light Toggle ---
        self.mode_toggle = QtWidgets.QToolButton()
        self.mode_toggle.setText("Dark")
        self.mode_toggle.setCheckable(True)
        self.mode_toggle.setChecked(True)
        self.mode_toggle.toggled.connect(self.toggle_mode)
        #toolbar.addWidget(self.mode_toggle) # we don't really need this shit (dark is fine).
        #toolbar.addSeparator()

        # --- Search Field ---
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("Search keyword…")
        self.search_edit.setMinimumWidth(150)
        self.search_edit.returnPressed.connect(self.on_search_return_pressed)
        self.search_edit.installEventFilter(self)
        toolbar.addWidget(self.search_edit)
        
        # --- Search Button (magnifier) ---
        self.search_btn = QtWidgets.QToolButton()
        self.search_btn.setText("🔍")
        self.search_btn.setToolTip("Search")
        self.search_btn.clicked.connect(self.apply_search_only)
        toolbar.addWidget(self.search_btn)

        # --- Case Sensitive Toggle ---
        self.case_sensitive_btn = QtWidgets.QToolButton()
        self.case_sensitive_btn.setText("Aa")
        self.case_sensitive_btn.setCheckable(True)
        self.case_sensitive_btn.setChecked(False)
        self.case_sensitive_btn.setToolTip("Case sensitive search (off)")
        self.case_sensitive_btn.toggled.connect(self.on_case_sensitive_toggled)

        # Add blue background when checked
        self.case_sensitive_btn.setStyleSheet("""
            QToolButton:checked {
                background-color: rgb(60, 100, 160);
                border: 1px solid rgb(80, 130, 200);
            }
        """)

        toolbar.addWidget(self.case_sensitive_btn)

        # --- Navigation Buttons ---
        self.btn_prev = QtWidgets.QToolButton()
        self.btn_prev.setArrowType(QtCore.Qt.ArrowType.LeftArrow)
        self.btn_prev.setToolTip("Previous match (Shift+Enter)")
        self.btn_prev.clicked.connect(self.go_prev_match)

        self.btn_next = QtWidgets.QToolButton()
        self.btn_next.setArrowType(QtCore.Qt.ArrowType.RightArrow)
        self.btn_next.setToolTip("Next match (Enter)")
        self.btn_next.clicked.connect(self.go_next_match)

        self.match_label = QtWidgets.QLabel("Matching cells: 0")
        self.match_label.setMinimumWidth(100)

        toolbar.addWidget(self.btn_prev)
        toolbar.addWidget(self.btn_next)
        toolbar.addWidget(self.match_label)

        toolbar.addSeparator()

        # --- Replace Section ---
        self.replace_container = QtWidgets.QWidget()
        repl_layout = QtWidgets.QHBoxLayout(self.replace_container)
        repl_layout.setContentsMargins(0, 0, 0, 0)
        self.replace_edit = QtWidgets.QLineEdit()
        self.replace_edit.setPlaceholderText("Replace with…")
        self.replace_edit.setMinimumWidth(150)
        self.replace_edit.textChanged.connect(self._update_replace_buttons)
        self.btn_replace_next = QtWidgets.QToolButton()
        self.btn_replace_next.setText("Replace Next")
        self.btn_replace_next.clicked.connect(self.on_replace_next)
        self.btn_replace_all = QtWidgets.QToolButton()
        self.btn_replace_all.setText("Replace All")
        self.btn_replace_all.clicked.connect(self.on_replace_all)
        repl_layout.addWidget(self.replace_edit)
        repl_layout.addWidget(self.btn_replace_next)
        repl_layout.addWidget(self.btn_replace_all)
        self.replace_container.setVisible(False)
        toolbar.addWidget(self.replace_container)

        toolbar.addSeparator()

        # --- Font Size Scaler ---
        font_container = QtWidgets.QWidget()
        font_layout = QtWidgets.QHBoxLayout(font_container)
        font_layout.setContentsMargins(0, 0, 0, 0)
        font_layout.setSpacing(2)

        self.font_decrease_btn = QtWidgets.QToolButton()
        self.font_decrease_btn.setText("−")
        self.font_decrease_btn.setToolTip("Decrease font size")
        self.font_decrease_btn.clicked.connect(lambda: self.change_font_size(-1))

        self.font_size_label = QtWidgets.QLabel("100%")
        self.font_size_label.setMinimumWidth(45)
        self.font_size_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        self.font_increase_btn = QtWidgets.QToolButton()
        self.font_increase_btn.setText("+")
        self.font_increase_btn.setToolTip("Increase font size")
        self.font_increase_btn.clicked.connect(lambda: self.change_font_size(1))

        font_layout.addWidget(self.font_decrease_btn)
        font_layout.addWidget(self.font_size_label)
        font_layout.addWidget(self.font_increase_btn)

        toolbar.addWidget(font_container)
        toolbar.addSeparator()

        # --- Other Actions ---
        self.btn_autosize = QtGui.QAction("Autosize", self)
        self.btn_autosize.triggered.connect(lambda: self.autosize_columns(force=True))
        toolbar.addAction(self.btn_autosize)

        # --- Table View ---
        self.table = SmoothScrollTableView()
        table_font = QtGui.QFont(self.table.font())
        table_font.setFamily(DEFAULT_DISPLAY_FONT_FAMILY)
        table_font.setPointSize(DEFAULT_DISPLAY_FONT_SIZE)
        self.table.setFont(table_font)
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Interactive)
        
        # === KEY CHANGE: Set vertical header to Fixed mode for manual management ===
        self.table.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Fixed)
        self.table.setWordWrap(True)
        
        # === NEW: Enable column dragging ===
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setDragEnabled(True)
        self.table.horizontalHeader().setDragDropMode(QtWidgets.QAbstractItemView.DragDropMode.InternalMove)
        self.table.horizontalHeader().sectionMoved.connect(self._on_column_moved)
        
        self.table.horizontalHeader().sectionClicked.connect(self._select_column)
        self.table.verticalHeader().sectionClicked.connect(self._select_row)
        self.table.horizontalHeader().setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(self._show_column_header_menu)
        self.table.verticalHeader().setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.verticalHeader().customContextMenuRequested.connect(self._show_row_header_menu)
        
        
                # === NEW: Visual feedback for column dragging ===
        self.table.horizontalHeader().setDragDropOverwriteMode(False)
        self.table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.table.horizontalHeader().sectionDoubleClicked.connect(self._rename_column)
        self.table.clicked.connect(self._handle_plus_column_click)
        self.table.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_cell_qa_menu)
        self._apply_scrollbar_style()
        
        # === NEW: Debounced row resizing system ===
        # Timer to debounce rapid column resize events
        self._row_resize_timer = QtCore.QTimer()
        self._row_resize_timer.setSingleShot(True)
        self._row_resize_timer.setInterval(120)  # 120ms debounce for responsiveness
        self._row_resize_timer.timeout.connect(self._resize_visible_rows)
        
        # Track which columns were resized since last update
        self._columns_resized = set()
        
        # Connect column resize signal
        self.table.horizontalHeader().sectionResized.connect(self._on_column_resized)
        
        # === NEW: Track scroll position for on-demand row sizing ===
        self._last_scroll_value = 0
        self.table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        
        # === NEW: Handle cell edits ===
        # This will be connected after model is created
        
        # --- Drop Zone ---
        self.dd = DragDropWidget()
        self.dd.fileDropped.connect(self.load_path)
        
        # Style the drop zone better
        self.dd.setStyleSheet("""
            DragDropWidget {
                background-color: rgba(60, 60, 60, 100);
                border: 2px dashed rgba(100, 100, 100, 150);
                border-radius: 12px;
            }
            DragDropWidget:hover {
                background-color: rgba(70, 70, 70, 120);
                border-color: rgba(120, 120, 120, 200);
            }
        """)

        self.dd_outer_container = QtWidgets.QWidget()
        dd_outer_layout = QtWidgets.QVBoxLayout(self.dd_outer_container)
        dd_outer_layout.setContentsMargins(8, 0, 8, 0)
        dd_outer_layout.setSpacing(0)
        dd_outer_layout.addWidget(self.dd)

        self.connect_container = QtWidgets.QWidget()
        connect_outer = QtWidgets.QVBoxLayout(self.connect_container)
        connect_outer.setContentsMargins(12, 8, 12, 10)
        connect_outer.setSpacing(8)
        connect_outer.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop)

        self.connect_label = QtWidgets.QLabel("Connect to Host")
        self.connect_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.connect_label.setStyleSheet("font-size: 14px; color: rgb(190, 190, 190);")
        connect_outer.addWidget(self.connect_label)

        connect_row = QtWidgets.QWidget()
        connect_layout = QtWidgets.QHBoxLayout(connect_row)
        connect_layout.setContentsMargins(0, 0, 0, 0)
        connect_layout.setSpacing(6)

        self.connect_input = QtWidgets.QLineEdit()
        self.connect_input.setPlaceholderText(f"{COLLAB_LINK_SCHEME}://host:port/token")
        self.connect_input.setMinimumWidth(320)
        self.connect_input.returnPressed.connect(self._on_connect_submit)

        self.connect_btn = QtWidgets.QPushButton("Connect")
        self.connect_btn.clicked.connect(self._on_connect_submit)

        connect_layout.addWidget(self.connect_input, 1)
        connect_layout.addWidget(self.connect_btn)
        connect_outer.addWidget(connect_row)

        # Status bar
        self.status = QtWidgets.QStatusBar()
        self.status.setFixedHeight(24)

        self.size_label = QtWidgets.QLabel("")
        self.size_label.setStyleSheet("font-weight: 500; padding-right: 10px;")
        self.status.addPermanentWidget(self.size_label)

        # Background task indicator (bottom-left corner)
        self.bg_indicator = QtWidgets.QLabel("")
        self.bg_indicator.setStyleSheet("""
            QLabel {
                background-color: #FFC107;
                color: black;
                padding: 4px 12px;
                border-radius: 4px;
                font-weight: 600;
                border: 1px solid #E65100;
            }
        """)
        self.bg_indicator.hide()
        self.status.addWidget(self.bg_indicator)  # Left side of status bar

        self.dirty_indicator = QtWidgets.QLabel("")
        self.dirty_indicator.setStyleSheet("""
            QLabel {
                color: #FFC107;
                font-weight: 600;
                padding: 2px 8px;
            }
        """)
        self.dirty_indicator.hide()
        self.status.addPermanentWidget(self.dirty_indicator)  # Right side of status bar

        self.edit_meta_indicator = QtWidgets.QLabel("")
        self.edit_meta_indicator.setStyleSheet(
            """
            QLabel {
                color: rgb(180, 180, 180);
                padding: 2px 8px;
            }
            """
        )
        self.edit_meta_indicator.setMinimumWidth(260)
        self.edit_meta_indicator.hide()
        self.status.addPermanentWidget(self.edit_meta_indicator)  # Lower-right style metadata

        # Timer for updating elapsed time & progress
        self.bg_timer = QtCore.QTimer()
        self.bg_timer.setInterval(500)  # Update every 500ms for smoother progress
        self.bg_timer.timeout.connect(self._update_bg_indicator)

        # Progress tracking
        self.bg_start_time = None
        self.bg_task_name = ""
        self.bg_total_rows = 0
        self.bg_current_row = 0

        # Add everything to layout
        outer.addWidget(self.controls)
        outer.addWidget(self.toolbar)
        outer.addWidget(self.dd_outer_container, 1)  # Keep small horizontal margin
        outer.addWidget(self.connect_container)
        outer.addWidget(self.table, 1)  # Give it stretch factor
        outer.addWidget(self.status)
        
        self.table.hide()
        self.toolbar.hide()
        self.status.hide()
        self._apply_session_mode_ui()
        
    def refresh_dims_display(self):
        """Updates the status bar with current row and column counts in bold."""
        if self.model:
            rows = self.model.df.shape[0]
            cols = self.model.df.shape[1]
            # Using HTML <b> tags to bold the specific numbers
            self.size_label.setText(f"| <b>{rows}</b> rows x <b>{cols}</b> columns |")
        else:
            self.size_label.clear()

    def _is_guest_mode(self) -> bool:
        return self._session_role == "guest"

    def _on_connect_submit(self):
        self._on_connect_requested(self.connect_input.text().strip())

    def _set_connect_controls_enabled(self, enabled: bool):
        self.connect_input.setEnabled(enabled)
        self.connect_btn.setEnabled(enabled)

    def _show_collab_notice(self, text: str, timeout_ms: int = 3800):
        if not text:
            return
        self._collab_last_notice_ts = time.time()
        self.status.showMessage(text, timeout_ms)
        try:
            local_pos = QtCore.QPoint(max(8, self.controls.width() - 240), self.controls.height() - 2)
            global_pos = self.controls.mapToGlobal(local_pos)
            QtWidgets.QToolTip.showText(global_pos, text, self.controls, self.controls.rect(), timeout_ms)
        except Exception:
            pass

    def _bind_table_selection_model(self):
        if self._table_selection_model is not None:
            try:
                self._table_selection_model.currentChanged.disconnect(self._on_table_current_changed)
            except Exception:
                pass
            try:
                self._table_selection_model.selectionChanged.disconnect(self._on_table_selection_changed)
            except Exception:
                pass

        self._table_selection_model = self.table.selectionModel()
        if self._table_selection_model is None:
            self._update_edit_meta_indicator()
            return

        self._table_selection_model.currentChanged.connect(self._on_table_current_changed)
        self._table_selection_model.selectionChanged.connect(self._on_table_selection_changed)
        self._update_edit_meta_indicator()

    def _on_table_current_changed(self, _current: QtCore.QModelIndex, _previous: QtCore.QModelIndex):
        self._update_edit_meta_indicator()
        self._queue_collab_presence_broadcast()

    def _on_table_selection_changed(self, _selected: QtCore.QItemSelection, _deselected: QtCore.QItemSelection):
        self._update_edit_meta_indicator()
        self._queue_collab_presence_broadcast()

    def _local_selected_cell(self) -> t.Optional[tuple[int, int]]:
        if not self.model:
            return None

        idx = self.table.currentIndex()
        if not idx.isValid():
            sel_model = self.table.selectionModel()
            if sel_model is not None:
                selected = sel_model.selectedIndexes()
                if selected:
                    idx = selected[0]

        if not idx.isValid():
            return None

        row = int(idx.row())
        col = int(idx.column())
        if row < 0 or col < 0:
            return None
        if row >= self.model.data_row_count() or col >= self.model.data_column_count():
            return None
        return row, col

    def _selected_edit_metadata_text(self) -> str:
        if not self.model:
            return ""
        pos = self._local_selected_cell()
        if pos is None:
            return ""
        metadata = self.model.cell_edit_metadata(pos[0], pos[1])
        if metadata is None:
            return ""
        edited_by, edited_at = metadata
        return f"Edited by {edited_by} at {edited_at}"

    def _update_edit_meta_indicator(self):
        text = self._selected_edit_metadata_text()
        if text:
            self.edit_meta_indicator.setText(text)
            self.edit_meta_indicator.show()
        else:
            self.edit_meta_indicator.hide()

    def _sync_remote_selection_overlay(self):
        self.table.set_remote_selections(dict(self._collab_remote_selections))

    def _clear_remote_selections(self):
        self._collab_remote_selections.clear()
        self.table.clear_remote_selections()

    def _queue_collab_presence_broadcast(self, force: bool = False):
        if self._collab_is_applying_remote_snapshot:
            return
        if not self.model:
            return

        selected_cell = self._local_selected_cell()
        if (not force) and selected_cell == self._collab_last_sent_selection:
            return
        self._collab_last_sent_selection = selected_cell

        payload = {
            "username": self._local_username,
            "row": -1,
            "col": -1,
        }
        if selected_cell is not None:
            payload["row"] = int(selected_cell[0])
            payload["col"] = int(selected_cell[1])

        if self._session_role == "host" and self._collab_server is not None:
            self._collab_server.broadcast_presence(payload)
            return

        if (
            self._session_role == "guest"
            and self._collab_client is not None
            and self._collab_client_connected
        ):
            self._collab_client.send_presence(payload)

    def _apply_collab_presence_payload(self, payload: dict[str, t.Any]) -> bool:
        if not isinstance(payload, dict):
            return False

        username = _normalize_username(payload.get("username"))
        if not username or username == self._local_username:
            return False

        try:
            row = int(payload.get("row", -1))
            col = int(payload.get("col", -1))
        except Exception:
            row, col = -1, -1

        if row < 0 or col < 0:
            if username in self._collab_remote_selections:
                self._collab_remote_selections.pop(username, None)
                self._sync_remote_selection_overlay()
            return True

        self._collab_remote_selections[username] = (row, col)
        self._sync_remote_selection_overlay()
        return True

    def _apply_session_mode_ui(self):
        has_model = self.model is not None
        is_guest = self._is_guest_mode()

        self.controls.back_btn.setVisible(has_model)
        self.controls.back_btn.setEnabled(has_model)

        show_share = has_model and not is_guest
        guest_count = max(0, int(self._collab_connected_guest_count))
        total_people = 1 + guest_count

        if self._session_role == "host" and guest_count > 0:
            self.controls.share_btn.setText(f"\u2197 Share \u2022 {total_people}")
        else:
            self.controls.share_btn.setText("\u2197 Share")

        self.controls.share_btn.setVisible(show_share)
        self.controls.share_btn.setEnabled(show_share)
        self.controls.share_count_lbl.setVisible(show_share)
        self.controls.share_count_lbl.setText(str(guest_count))
        self.controls.share_count_lbl.setToolTip("Connected guests")
        if self._session_role == "host":
            self.controls.share_btn.setToolTip("Copy the current share link")
        elif is_guest:
            self.controls.share_btn.setToolTip("Guests cannot host")
        else:
            self.controls.share_btn.setToolTip("Start hosting and copy share link")

        self.btn_open.setEnabled(not is_guest)
        self.dd.set_drop_enabled(not is_guest)
        self.connect_container.setVisible(not has_model)
        self._set_connect_controls_enabled((not self._collab_connecting) and (not self._collab_client_connected))
        self.sheet_combo.setEnabled((not is_guest) and len(self._xlsx_sheet_names) > 1)
        if not has_model:
            self.edit_meta_indicator.hide()
        self.controls.sync_title_balance()

    def _attach_collab_model_signals(self, model: DataFrameModel):
        model.dataChanged.connect(self._on_collab_model_mutation)
        model.rowsInserted.connect(self._on_collab_model_mutation)
        model.rowsRemoved.connect(self._on_collab_model_mutation)
        model.columnsInserted.connect(self._on_collab_model_mutation)
        model.columnsRemoved.connect(self._on_collab_model_mutation)
        model.layoutChanged.connect(self._on_collab_model_mutation)
        model.headerDataChanged.connect(self._on_collab_model_mutation)

    def _on_collab_model_mutation(self, *args):
        if self._collab_is_applying_remote_snapshot:
            return
        self._queue_collab_snapshot_broadcast(immediate=False)

    def _queue_collab_snapshot_broadcast(self, immediate: bool = False):
        if self._collab_is_applying_remote_snapshot:
            return
        if not self.model:
            return
        if self._session_role == "host":
            if self._collab_server is None:
                return
        elif self._session_role == "guest":
            if self._collab_client is None or not self._collab_client_connected:
                return
        else:
            return

        if immediate:
            self._collab_broadcast_timer.stop()
            self._broadcast_current_snapshot()
        else:
            self._collab_broadcast_timer.start()

    def _build_collab_snapshot(self) -> t.Optional[dict[str, t.Any]]:
        if not self.model:
            return None

        columns = [str(col) for col in list(self.model.df.columns)]
        rows: list[list[t.Any]] = []
        for raw_row in self.model.df.values.tolist():
            rows.append([_json_safe_value(value) for value in raw_row])

        sheet_names = list(self._xlsx_sheet_names)
        if not sheet_names and self.model.xlsx_sheet_name:
            sheet_names = [str(self.model.xlsx_sheet_name)]

        display_name = self.current_path.name if self.current_path else self.controls.title_lbl.text().strip()
        if not display_name:
            display_name = "Shared data"

        snapshot = {
            "protocol": COLLAB_PROTOCOL_VERSION,
            "revision": int(self._collab_revision),
            "display_name": display_name,
            "source_format": self.model.source_format or ".csv",
            "sheet_name": self.model.xlsx_sheet_name or "Sheet1",
            "sheet_names": sheet_names,
            "columns": columns,
            "rows": rows,
            "cell_styles": _serialize_cell_styles(self.model.cell_styles()),
            "is_dirty": bool(self.model.is_dirty),
        }
        return snapshot

    def _broadcast_current_snapshot(self):
        if self._collab_is_applying_remote_snapshot:
            return

        snapshot = self._build_collab_snapshot()
        if not snapshot:
            return

        if self._session_role == "host" and self._collab_server is not None:
            self._collab_revision += 1
            snapshot["revision"] = int(self._collab_revision)
            self._collab_server.broadcast_snapshot(snapshot)
            return

        if (
            self._session_role == "guest"
            and self._collab_client is not None
            and self._collab_client_connected
        ):
            snapshot["revision"] = int(self._collab_revision)
            self._collab_client.send_snapshot(snapshot)

    def _install_dataframe_model(
        self,
        df: pd.DataFrame,
        *,
        cell_styles: dict[tuple[int, int], dict[str, t.Any]],
        source_format: str,
        xlsx_sheet_name: t.Optional[str],
        xlsx_base_font,
        loaded_sheet_names: list[str],
        title: str,
        status_message: str,
        current_path: t.Optional[Path],
        mark_dirty: bool = False,
    ):
        self.current_path = current_path
        self.controls.title_lbl.setText(title)
        self.window().setWindowTitle(title)

        model = DataFrameModel(
            df,
            cell_styles=cell_styles,
            source_format=source_format,
            xlsx_sheet_name=xlsx_sheet_name or "Sheet1",
            xlsx_base_font=xlsx_base_font,
        )
        model.set_edit_actor(self._local_username)
        model._is_dark = self.mode_toggle.isChecked()
        model.dataChangedHard.connect(self.table.viewport().update)
        model.matchesChanged.connect(self._on_matches_changed)
        model.dirtyStateChanged.connect(self.update_dirty_indicator)
        model.dataChanged.connect(self._on_cell_data_changed)
        self._attach_collab_model_signals(model)
        self.model = model

        self.table.setModel(self.model)
        self._bind_table_selection_model()
        self._sync_remote_selection_overlay()
        self.refresh_dims_display()
        self.table.setItemDelegate(HighlightDelegate(self.model))

        if source_format in [".xlsx", ".xls"]:
            self._set_sheet_selector(loaded_sheet_names, xlsx_sheet_name)
        else:
            self._set_sheet_selector([], None)

        if self.model._is_dark:
            self.table.setStyleSheet("QTableView { gridline-color: rgb(80, 80, 80); }")
            self._apply_scrollbar_style()
        else:
            self.table.setStyleSheet("QTableView { gridline-color: rgb(0, 0, 0); }")
            self._apply_scrollbar_style()

        self.table.show()
        self.dd_outer_container.hide()
        self.connect_container.hide()
        self.toolbar.show()
        self.status.show()
        self.controls.back_btn.setEnabled(True)
        self.status.showMessage(status_message)

        self.autosize_columns(force=True)
        QtCore.QTimer.singleShot(50, self._autosize_all_rows)

        self.current_match_pos = -1
        self.base_font_size = 10
        self.font_scale = 1.0
        self.model.set_search_term("")
        self._on_matches_changed(0)

        self.model.is_dirty = bool(mark_dirty)
        self.update_dirty_indicator()
        self._update_edit_meta_indicator()
        self._apply_session_mode_ui()
        self.window().showMaximized()
        QtCore.QTimer.singleShot(0, lambda: self._queue_collab_presence_broadcast(force=True))

    def _apply_collab_snapshot(self, snapshot: dict[str, t.Any], source_label: str) -> bool:
        if not isinstance(snapshot, dict):
            return False

        raw_columns = snapshot.get("columns")
        raw_rows = snapshot.get("rows")
        if not isinstance(raw_columns, list) or not isinstance(raw_rows, list):
            return False

        columns = [str(col) for col in raw_columns]
        normalized_rows: list[list[t.Any]] = []
        for row in raw_rows:
            values = list(row) if isinstance(row, list) else []
            if len(values) < len(columns):
                values.extend([None] * (len(columns) - len(values)))
            elif len(values) > len(columns):
                values = values[: len(columns)]
            normalized_rows.append(values)

        df = pd.DataFrame(normalized_rows, columns=columns)
        cell_styles = _deserialize_cell_styles(snapshot.get("cell_styles"))
        source_format = str(snapshot.get("source_format") or ".csv").lower()
        sheet_name = str(snapshot.get("sheet_name") or "Sheet1")

        raw_sheet_names = snapshot.get("sheet_names")
        if isinstance(raw_sheet_names, list):
            loaded_sheet_names = [str(name) for name in raw_sheet_names if str(name).strip()]
        else:
            loaded_sheet_names = []
        if not loaded_sheet_names:
            loaded_sheet_names = [sheet_name]

        if self._session_role == "host" and self.current_path is not None:
            display_name = self.current_path.name
        else:
            display_name = str(snapshot.get("display_name") or "Shared data")
        revision = snapshot.get("revision")
        if isinstance(revision, int):
            self._collab_revision = max(self._collab_revision, revision)

        self._collab_is_applying_remote_snapshot = True
        try:
            self._install_dataframe_model(
                df,
                cell_styles=cell_styles,
                source_format=source_format,
                xlsx_sheet_name=sheet_name,
                xlsx_base_font=None,
                loaded_sheet_names=loaded_sheet_names,
                title=display_name,
                status_message=f"Synced from {source_label}",
                current_path=self.current_path,
                mark_dirty=bool(snapshot.get("is_dirty", False)),
            )
        finally:
            self._collab_is_applying_remote_snapshot = False
        return True

    def _start_host_session(self) -> bool:
        if self._is_guest_mode():
            return False
        if not self.model:
            return False

        if self._collab_server is None:
            token = secrets.token_urlsafe(12)
            server = CollaborationServer(token, parent=self)
            server.snapshotReceived.connect(self._on_collab_server_snapshot_received)
            server.presenceReceived.connect(self._on_collab_server_presence_received)
            server.errorOccurred.connect(self._on_collab_server_error)
            server.clientCountChanged.connect(self._on_collab_server_client_count_changed)
            server.guestConnected.connect(self._on_collab_guest_connected)
            server.guestDisconnected.connect(self._on_collab_guest_disconnected)
            if not server.start(0):
                return False
            self._collab_server = server
            self._collab_session_token = token
            host_ip = _guess_local_ipv4()
            self._collab_share_link = _build_collab_link(host_ip, int(server.port or 0), token)

        self._session_role = "host"
        self._apply_session_mode_ui()
        self._queue_collab_snapshot_broadcast(immediate=True)
        self._queue_collab_presence_broadcast(force=True)
        return True

    def _on_share_requested(self):
        if self._is_guest_mode():
            QtWidgets.QMessageBox.information(
                self,
                "Guest Session",
                "Guests cannot host sessions. Disconnect first to host a file.",
            )
            return
        if not self.model:
            QtWidgets.QMessageBox.information(self, "Nothing to Share", "Open a file before sharing.")
            return
        if not self._start_host_session():
            QtWidgets.QMessageBox.warning(self, "Share Error", "Could not start host session.")
            return

        clipboard = QtWidgets.QApplication.clipboard()
        clipboard.setText(self._collab_share_link)
        self.status.showMessage("Share link copied to clipboard.", 3500)

    def _on_connect_requested(self, raw_link: str):
        parsed = _parse_collab_link(raw_link)
        if parsed is None:
            QtWidgets.QMessageBox.warning(
                self,
                "Invalid Link",
                f"Paste a valid link: {COLLAB_LINK_SCHEME}://host:port/token",
            )
            return

        if self._session_role == "host":
            reply = QtWidgets.QMessageBox.question(
                self,
                "Switch to Guest",
                "You are hosting right now. Stop hosting and connect as a guest?",
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
            )
            if reply != QtWidgets.QMessageBox.StandardButton.Yes:
                return

        if self.model and self._session_role != "guest":
            reply = QtWidgets.QMessageBox.question(
                self,
                "Replace Current Data",
                "Connecting to a host will replace your current local data. Continue?",
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
            )
            if reply != QtWidgets.QMessageBox.StandardButton.Yes:
                return

        self._stop_collaboration()
        if self.model is not None:
            self.table.setModel(None)
            self._bind_table_selection_model()
            self.model = None
            self.current_match_pos = -1
            self.refresh_dims_display()
            self._set_sheet_selector([], None)
            self.dirty_indicator.hide()
        self.current_path = None
        self.toolbar.hide()
        self.status.hide()
        self.table.hide()
        self.dd_outer_container.show()
        self.connect_container.show()
        self.controls.back_btn.setEnabled(False)

        host, port, token = parsed
        client = CollaborationClient(host, port, token, username=self._local_username, parent=self)
        client.connectedChanged.connect(self._on_collab_client_connected_changed)
        client.errorOccurred.connect(self._on_collab_client_error)
        client.snapshotReceived.connect(self._on_collab_client_snapshot_received)
        client.presenceReceived.connect(self._on_collab_client_presence_received)
        self._collab_client = client
        self._session_role = "guest"
        self._collab_client_connected = False
        self._collab_connecting = True
        self.dd.icon.setText("Connecting to host...")
        self._apply_session_mode_ui()
        client.start()

    def _on_collab_client_connected_changed(self, connected: bool):
        self._collab_client_connected = bool(connected)
        self._collab_connecting = False
        if connected:
            self._session_role = "guest"
            self._show_collab_notice(f"Connected as {self._local_username}. Waiting for shared file...", 4200)
            self._queue_collab_presence_broadcast(force=True)
        else:
            if self._session_role == "guest" and self._collab_client is not None:
                self._session_role = "local"
                self._set_connect_controls_enabled(True)
                if self.model is None:
                    self.dd.icon.setText("Drop a file or connect to a host")
            if self._collab_client is not None and not self._collab_client.is_connected():
                self._collab_client = None
            self._collab_last_sent_selection = None
            self._clear_remote_selections()
        self._apply_session_mode_ui()

    def _on_collab_client_error(self, message: str):
        self._collab_connecting = False
        if message:
            self.status.showMessage(message, 5000)
        if self._session_role == "guest" and self.model is None:
            self.dd.icon.setText("Connection failed. Paste a host link and retry.")
            self._set_connect_controls_enabled(True)
        self._apply_session_mode_ui()

    def _on_collab_server_error(self, message: str):
        if message:
            self.status.showMessage(message, 5000)

    def _on_collab_guest_connected(self, username: str, count: int):
        self._collab_connected_guest_count = max(0, int(count))
        self._apply_session_mode_ui()
        if self._session_role == "host":
            safe_username = _normalize_username(username)
            self._show_collab_notice(f"{safe_username} has connected!", 4200)
            self._queue_collab_snapshot_broadcast(immediate=True)
            self._queue_collab_presence_broadcast(force=True)

    def _on_collab_guest_disconnected(self, username: str, count: int):
        self._collab_connected_guest_count = max(0, int(count))
        safe_username = _normalize_username(username)
        self._collab_remote_selections.pop(safe_username, None)
        self._sync_remote_selection_overlay()
        self._apply_session_mode_ui()
        if self._session_role == "host":
            self._show_collab_notice(f"{safe_username} has disconnected.", 3600)
            if self._collab_server is not None:
                self._collab_server.broadcast_presence(
                    {
                        "username": safe_username,
                        "row": -1,
                        "col": -1,
                    }
                )

    def _on_collab_server_client_count_changed(self, count: int):
        previous_count = int(self._collab_connected_guest_count)
        self._collab_connected_guest_count = max(0, int(count))
        self._apply_session_mode_ui()
        if self._session_role == "host":
            # Fallback in case guest name signal was not received.
            now = time.time()
            if count > previous_count and (now - self._collab_last_notice_ts) > 0.8:
                self._show_collab_notice(f"A guest has connected. {count} online.", 3200)
            elif count < previous_count and (now - self._collab_last_notice_ts) > 0.8:
                self._show_collab_notice(f"A guest disconnected. {count} online.", 3200)

    def _on_collab_server_presence_received(self, payload: dict[str, t.Any]):
        if self._session_role != "host":
            return
        if not isinstance(payload, dict):
            return
        if self._apply_collab_presence_payload(payload):
            if self._collab_server is not None:
                self._collab_server.broadcast_presence(payload)

    def _on_collab_client_presence_received(self, payload: dict[str, t.Any]):
        if self._session_role != "guest":
            return
        self._apply_collab_presence_payload(payload)

    def _on_collab_server_snapshot_received(self, payload: dict[str, t.Any]):
        if self._session_role != "host":
            return
        if self._apply_collab_snapshot(payload, "guest"):
            self._queue_collab_snapshot_broadcast(immediate=True)

    def _on_collab_client_snapshot_received(self, payload: dict[str, t.Any]):
        if self._session_role != "guest":
            return
        self._apply_collab_snapshot(payload, "host")

    def _stop_collaboration(self):
        self._collab_broadcast_timer.stop()

        if self._collab_server is not None:
            try:
                self._collab_server.snapshotReceived.disconnect(self._on_collab_server_snapshot_received)
            except Exception:
                pass
            try:
                self._collab_server.presenceReceived.disconnect(self._on_collab_server_presence_received)
            except Exception:
                pass
            try:
                self._collab_server.errorOccurred.disconnect(self._on_collab_server_error)
            except Exception:
                pass
            try:
                self._collab_server.clientCountChanged.disconnect(self._on_collab_server_client_count_changed)
            except Exception:
                pass
            try:
                self._collab_server.guestConnected.disconnect(self._on_collab_guest_connected)
            except Exception:
                pass
            try:
                self._collab_server.guestDisconnected.disconnect(self._on_collab_guest_disconnected)
            except Exception:
                pass
            self._collab_server.stop()
            self._collab_server = None

        if self._collab_client is not None:
            if self._session_role == "guest" and self._collab_client_connected:
                try:
                    self._collab_client.send_presence(
                        {
                            "username": self._local_username,
                            "row": -1,
                            "col": -1,
                        }
                    )
                except Exception:
                    pass
            try:
                self._collab_client.connectedChanged.disconnect(self._on_collab_client_connected_changed)
            except Exception:
                pass
            try:
                self._collab_client.errorOccurred.disconnect(self._on_collab_client_error)
            except Exception:
                pass
            try:
                self._collab_client.snapshotReceived.disconnect(self._on_collab_client_snapshot_received)
            except Exception:
                pass
            try:
                self._collab_client.presenceReceived.disconnect(self._on_collab_client_presence_received)
            except Exception:
                pass
            self._collab_client.stop()
            self._collab_client = None

        self._collab_client_connected = False
        self._collab_connecting = False
        self._collab_last_sent_selection = None
        self._collab_connected_guest_count = 0
        self._clear_remote_selections()
        self._collab_session_token = ""
        self._collab_share_link = ""
        self._collab_revision = 0
        self._session_role = "local"
        self._apply_session_mode_ui()

    def shutdown(self):
        self._stop_collaboration()

    def _set_sheet_selector(self, sheet_names: list[str], selected_sheet: t.Optional[str]):
        self._xlsx_sheet_names = sheet_names
        has_any = len(sheet_names) > 0
        self.sheet_label.setVisible(True)
        self.sheet_combo.setVisible(True)
        self.sheet_combo.setEnabled((not self._is_guest_mode()) and has_any and len(sheet_names) > 1)

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        if has_any:
            self.sheet_combo.addItems(sheet_names)
            if selected_sheet and selected_sheet in sheet_names:
                self.sheet_combo.setCurrentText(selected_sheet)
            elif sheet_names:
                self.sheet_combo.setCurrentIndex(0)
        else:
            self.sheet_combo.addItem("(No sheet)")
            self.sheet_combo.setCurrentIndex(0)
        self.sheet_combo.blockSignals(False)

    def _on_sheet_changed(self, sheet_name: str):
        if self._is_guest_mode():
            return
        if self._sheet_reload_in_progress:
            return
        if not sheet_name or not self.current_path or self.current_path.suffix.lower() not in [".xlsx", ".xls"]:
            return
        if self.model and self.model.xlsx_sheet_name == sheet_name:
            return

        if self.model and self.model.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Save before switching sheets?",
                (QtWidgets.QMessageBox.StandardButton.Yes |
                QtWidgets.QMessageBox.StandardButton.No |
                QtWidgets.QMessageBox.StandardButton.Cancel),
            )
            if reply == QtWidgets.QMessageBox.StandardButton.Cancel:
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.setCurrentText(self.model.xlsx_sheet_name if self.model else "")
                self.sheet_combo.blockSignals(False)
                return
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                if not self.save_file():
                    self.sheet_combo.blockSignals(True)
                    self.sheet_combo.setCurrentText(self.model.xlsx_sheet_name if self.model else "")
                    self.sheet_combo.blockSignals(False)
                    return

        self._sheet_reload_in_progress = True
        try:
            self.load_path(self.current_path, xlsx_sheet_name=sheet_name)
        finally:
            self._sheet_reload_in_progress = False
            
    def on_case_sensitive_toggled(self, checked: bool):
        """Toggle case sensitivity and refresh search"""
        if self.model:
            self.model.case_sensitive = checked
            self.case_sensitive_btn.setToolTip(f"Case sensitive search ({'on' if checked else 'off'})")
            
            # Re-run search with new setting
            search_text = self.search_edit.text()
            if search_text:
                current_term = self.model.search_term
                self.model.set_search_term("")
                self.model.set_search_term(current_term)
                self.table.viewport().update()  # 👈 Force immediate repaint
                
    def reset_viewer(self):
        """Clear all loaded data and return to drag-and-drop screen"""
        self._stop_collaboration()
        # Clear the model
        if self.model:
            self.table.setModel(None)
            self._bind_table_selection_model()
            self.model = None
        
        # Reset UI state
        self.current_path = None
        self.current_match_pos = -1
        self.search_edit.clear()
        self.replace_edit.clear()
        self.replace_container.setVisible(False)
        self._set_sheet_selector([], None)
        
        # Reset window title
        self.controls.title_lbl.setText(APP_NAME)
        self.window().setWindowTitle(APP_NAME)
        
        # Reset indicators
        self.match_label.setText("Matching cells: 0")
        self.status.clearMessage()
        self.size_label.clear()
        self.dirty_indicator.hide()
        self._hide_background_indicator()
        self.toolbar.hide()
        self.status.hide()
        
        # Show drag-and-drop, hide table
        self.table.hide()
        self.dd_outer_container.show()
        self.connect_container.show()
        self.controls.back_btn.setEnabled(False)
        self._set_connect_controls_enabled(True)
        self._apply_session_mode_ui()
        self.dd.icon.setText(DROP_HINT_TEXT)
        
               
    def _show_background_indicator(self, task_name: str, total_rows: int, current_row: int = 0):
        """Show the background processing indicator with progress and elapsed time"""
        self.bg_task_name = task_name
        self.bg_start_time = time.time()
        self.bg_total_rows = total_rows
        self.bg_current_row = current_row
        self.bg_indicator.show()
        self._update_bg_indicator()
        self.bg_timer.start()

    def _hide_background_indicator(self):
        """Hide the background processing indicator"""
        self.bg_timer.stop()
        self.bg_indicator.hide()
        self.bg_start_time = None
        self.bg_task_name = ""
        self.bg_total_rows = 0
        self.bg_current_row = 0

    def _update_bg_indicator(self):
        """Update the progress and elapsed time in the indicator"""
        if self.bg_start_time is None or not self.isVisible():
            return
        
        elapsed = int(time.time() - self.bg_start_time)
        
        # Show both progress % and elapsed time
        if self.bg_total_rows > 0:
            progress = (self.bg_current_row / self.bg_total_rows) * 100
            self.bg_indicator.setText(f"{self.bg_task_name}… ({progress:.1f}% - {elapsed}s)")
        else:
            self.bg_indicator.setText(f"{self.bg_task_name}… ({elapsed}s)")
    
    def _on_column_moved(self, logical_index, old_visual_index, new_visual_index):
        """Reorder DataFrame columns when user drags column header"""
        if not self.model or old_visual_index == new_visual_index:
            return
        if logical_index >= len(self.model.df.columns):
            return
        
        try:
            # Get current visual order
            header = self.table.horizontalHeader()
            visual_order = [header.logicalIndex(i) for i in range(len(self.model.df.columns))]
            
            # Use layout change signals for efficient update (no flicker)
            self.model.layoutAboutToBeChanged.emit()
            
            # Reorder DataFrame to match visual order
            new_columns = [self.model.df.columns[i] for i in visual_order]
            self.model.df = self.model.df[new_columns]
            self.model.reorder_cell_styles(visual_order)
            self.model.is_dirty = True
            self.model.layoutChanged.emit()
            
            # Reapply search highlights (column indices changed)
            search_text = self.search_edit.text()
            if search_text:
                self.model.set_search_term("")
                self.model.set_search_term(search_text)
            
            # Resize rows (column widths are preserved by the header)
            self._autosize_all_rows()
            
            # Update status
            col_name = new_columns[new_visual_index]
            self.status.showMessage(f"Moved column '{col_name}' to position {new_visual_index + 1}")
            
        except Exception as e:
            print(f"Error moving column: {e}")
            self.status.showMessage("Error moving column")
            
    def _toggle_max_restore(self):
        if self.window().isMaximized() or self.window().isFullScreen():
            self.window().showNormal()
        else:
            self.window().showMaximized()
            
    def _apply_scrollbar_style(self):
        """Apply custom scrollbar styling with gray sliders"""
        scrollbar_style = """
            QScrollBar:vertical {
                border: none;
                background: rgb(30, 30, 30);
                width: 14px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: rgb(120, 120, 120);
                min-height: 30px;
                border-radius: 7px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgb(140, 140, 140);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
            
            QScrollBar:horizontal {
                border: none;
                background: rgb(30, 30, 30);
                height: 14px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: rgb(120, 120, 120);
                min-width: 30px;
                border-radius: 7px;
            }
            QScrollBar::handle:horizontal:hover {
                background: rgb(140, 140, 140);
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                background: none;
            }
        """
        self.table.setStyleSheet(self.table.styleSheet() + scrollbar_style)
                             
    def on_search_return_pressed(self):
        search_text = self.search_edit.text()
        if not self.model:
            return
            
        current_model_term = self.model.search_term or ""
        
        # CHECK: Is this a new search term?
        if search_text != current_model_term:
            # --- CASE 1: New Term (First Enter) ---
            
            # A. Capture current scroll position (Pixel-perfect)
            current_v_scroll = self.table.verticalScrollBar().value()
            
            # B. Update the Search (Highlights appear)
            self.model.set_search_term(search_text)
            
            # C. Show Replace bar
            self.replace_container.setVisible(bool(search_text.strip()))
            self._update_replace_buttons()
            
            # D. FORCE INSTANT RESTORE (No animation, no jumping)
            sb = self.table.verticalScrollBar()
            if hasattr(sb, 'setValueInstant'):
                sb.setValueInstant(current_v_scroll)
            else:
                sb.setValue(current_v_scroll)
            
            # E. Smart Start: Set to -1 so next press goes to first match
            self.current_match_pos = -1
        
        # Always navigate to next match when Enter is pressed
        self.go_next_match()
     
    def apply_search_only(self):
        """Updates highlights based on search text without moving the view."""
        if not self.model:
            return
            
        search_text = self.search_edit.text()
        
        # Update the model's search term to trigger repainting
        self.model.set_search_term(search_text)
        
        # Show/Hide replace container based on input
        self.replace_container.setVisible(bool(search_text.strip()))
        self._update_replace_buttons()
        
        # Reset match position but DON'T scroll
        self.current_match_pos = -1
        self.table.viewport().update()
        self.status.showMessage(f"Search updated. Found {self.model.total_matches()} matches.", 2000)
              
    def _prompt_add_column(self):
        if not self.model:
            return
        new_name, ok = QtWidgets.QInputDialog.getText(
            self, "New Column", "Enter column name:"
        )
        if ok and new_name.strip():
            self.model.add_new_column(new_name.strip())
            self.refresh_dims_display()
            self.autosize_columns(force=True)
            self._autosize_all_rows()
            self.status.showMessage(f"Added column: {new_name}")

    def _add_new_row(self):
        if not self.model:
            return
        self.model.add_new_row()
        self.refresh_dims_display()
        self._autosize_all_rows()
        new_row = self.model.data_row_count() - 1
        if new_row >= 0:
            self.table.scrollTo(self.model.index(new_row, 0), QtWidgets.QAbstractItemView.ScrollHint.PositionAtBottom)
        self.status.showMessage("Added row", 2000)

    def _remove_columns(self, col_indices: list[int]) -> bool:
        if not self.model:
            return False
        valid_indices = sorted({i for i in col_indices if 0 <= i < len(self.model._df.columns)})
        if not valid_indices:
            return False

        msg = f"Delete {len(valid_indices)} column(s)?"
        reply = QtWidgets.QMessageBox.question(
            self,
            "Delete Columns",
            msg,
            QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
        )
        if reply != QtWidgets.QMessageBox.StandardButton.Yes:
            return False

        self.model.layoutAboutToBeChanged.emit()
        cols_to_drop = [self.model._df.columns[i] for i in valid_indices]
        self.model._df.drop(columns=cols_to_drop, inplace=True)
        self.model.drop_style_columns(valid_indices)
        self.model.is_dirty = True
        self.model.layoutChanged.emit()
        self.refresh_dims_display()
        return True

    def _remove_rows(self, row_indices: list[int]) -> bool:
        if not self.model:
            return False
        valid_indices = sorted({i for i in row_indices if 0 <= i < len(self.model._df.index)})
        if not valid_indices:
            return False

        msg = f"Delete {len(valid_indices)} row(s)?"
        reply = QtWidgets.QMessageBox.question(
            self,
            "Delete Rows",
            msg,
            QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
        )
        if reply != QtWidgets.QMessageBox.StandardButton.Yes:
            return False

        self.model.layoutAboutToBeChanged.emit()
        self.model._df.drop(self.model._df.index[valid_indices], inplace=True)
        self.model._df.reset_index(drop=True, inplace=True)
        self.model.drop_style_rows(valid_indices)
        self.model.is_dirty = True
        self.model.layoutChanged.emit()
        self.refresh_dims_display()
        if self.model.search_term:
            self.model._rebuild_matches()
        return True

    def _show_column_header_menu(self, pos: QtCore.QPoint):
        if not self.model:
            return
        header = self.table.horizontalHeader()
        logical_index = header.logicalIndexAt(pos)
        if logical_index < 0:
            return

        menu = QtWidgets.QMenu(self)
        if logical_index == len(self.model.df.columns):
            add_action = menu.addAction("Add Column")
            chosen = menu.exec(header.mapToGlobal(pos))
            if chosen == add_action:
                self._prompt_add_column()
            return

        self.table.selectColumn(logical_index)
        rename_action = menu.addAction("Rename Column")
        remove_action = menu.addAction("Remove Column")
        chosen = menu.exec(header.mapToGlobal(pos))
        if chosen == rename_action:
            self._rename_column(logical_index)
        elif chosen == remove_action:
            self._remove_columns([logical_index])

    def _show_row_header_menu(self, pos: QtCore.QPoint):
        if not self.model:
            return
        header = self.table.verticalHeader()
        logical_index = header.logicalIndexAt(pos)
        if logical_index < 0:
            return

        menu = QtWidgets.QMenu(self)
        if logical_index == len(self.model.df.index):
            add_action = menu.addAction("Add Row")
            chosen = menu.exec(header.mapToGlobal(pos))
            if chosen == add_action:
                self._add_new_row()
            return

        self.table.selectRow(logical_index)
        remove_action = menu.addAction("Remove Row")
        chosen = menu.exec(header.mapToGlobal(pos))
        if chosen == remove_action:
            self._remove_rows([logical_index])

    def _select_column(self, logical_index: int):
        """Select entire column when header is clicked."""
        if not self.model:
            return
        if logical_index == len(self.model.df.columns):
            self._prompt_add_column()
            return
        self.table.selectColumn(logical_index)

    def _select_row(self, logical_index: int):
        """Select entire row when row header is clicked."""
        if not self.model:
            return
        if logical_index == len(self.model.df.index):
            self._add_new_row()
            return
        self.table.selectRow(logical_index)
    
    
    def load_path(self, path: Path, xlsx_sheet_name: t.Optional[str] = None):
        if self._is_guest_mode():
            QtWidgets.QMessageBox.information(
                self,
                "Guest Session",
                "Guests cannot open local files while connected to a host.",
            )
            return

        self._hide_background_indicator()
        self.dd.icon.setText("Loading...")
        QtWidgets.QApplication.processEvents()
        cell_styles: dict[tuple[int, int], dict[str, t.Any]] = {}
        loaded_sheet_names: list[str] = []
        selected_sheet_name = xlsx_sheet_name
        xlsx_base_font = None
        
        try:
            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path, low_memory=False)
            elif path.suffix.lower() in [".xls", ".xlsx"]:
                if path.suffix.lower() == ".xlsx":
                    df, cell_styles, selected_sheet_name, loaded_sheet_names, xlsx_base_font = load_xlsx_dataframe_with_styles(
                        path,
                        selected_sheet_name,
                    )
                    if not loaded_sheet_names:
                        loaded_sheet_names = [selected_sheet_name]
                else:
                    try:
                        excel_file = pd.ExcelFile(path)
                        loaded_sheet_names = list(excel_file.sheet_names)
                        if loaded_sheet_names:
                            if selected_sheet_name not in loaded_sheet_names:
                                selected_sheet_name = loaded_sheet_names[0]
                        else:
                            selected_sheet_name = selected_sheet_name or "Sheet1"
                        df = pd.read_excel(path, sheet_name=selected_sheet_name)
                    except Exception:
                        df = pd.read_excel(path)
                        selected_sheet_name = selected_sheet_name or "Sheet1"
                        loaded_sheet_names = [selected_sheet_name]
            elif path.suffix.lower() == ".jsonl":
                df = pd.read_json(path, lines=True)
            else:
                QtWidgets.QMessageBox.warning(self, "Unsupported", f"Unsupported file: {path}")
                return
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to open file:\n{e}")
            return

        filename = path.name
        if path.suffix.lower() in [".xlsx", ".xls"]:
            status_message = (
                f"Loaded: {path}  |  {df.shape[0]} rows x {df.shape[1]} cols  |  Sheet: {selected_sheet_name}"
            )
        else:
            status_message = f"Loaded: {path}  |  {df.shape[0]} rows x {df.shape[1]} cols"

        self._install_dataframe_model(
            df,
            cell_styles=cell_styles,
            source_format=path.suffix.lower(),
            xlsx_sheet_name=selected_sheet_name,
            xlsx_base_font=xlsx_base_font,
            loaded_sheet_names=loaded_sheet_names,
            title=filename,
            status_message=status_message,
            current_path=path,
            mark_dirty=False,
        )

        if self._session_role == "host":
            self._queue_collab_snapshot_broadcast(immediate=True)

    def open_file_dialog(self):
        """Open new file with unsaved changes check"""
        if self._is_guest_mode():
            QtWidgets.QMessageBox.information(
                self,
                "Guest Session",
                "Guests cannot open local files while connected to a host.",
            )
            return

        # Check dirty state BEFORE closing
        if self.model and self.model.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before opening a new file?",
                (QtWidgets.QMessageBox.StandardButton.Yes | 
                QtWidgets.QMessageBox.StandardButton.No | 
                QtWidgets.QMessageBox.StandardButton.Cancel)
            )

            if reply == QtWidgets.QMessageBox.StandardButton.Cancel:
                return  # Abort - keep current file open
            elif reply == QtWidgets.QMessageBox.StandardButton.Yes:
                if not self.save_file():
                    return  # Abort if save failed
        
        # Only proceed if user didn't cancel
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Open File", str(Path.home()), "Data Files (*.csv *.xls *.xlsx *.jsonl)"
        )
        if path:
            self.load_path(Path(path))
            
    def change_font_size(self, delta: int):
        # Adjust scale by 10% per step
        self.font_scale += delta * 0.1
        self.font_scale = max(0.5, min(2.0, self.font_scale))  # 50% to 200%
        
        # Update label
        self.font_size_label.setText(f"{int(self.font_scale * 100)}%")
        
        # Apply new font size to table
        new_size = int(self.base_font_size * self.font_scale)
        font = self.table.font()
        font.setPointSize(new_size)
        self.table.setFont(font)
        
        # Refresh table to apply changes
        self.table.viewport().update()
        self._autosize_all_rows()  # Recalculate all rows after font change
        
    def save_file(self):
        if not self.model:
            QtWidgets.QMessageBox.information(self, "Nothing to save", "Open a file first.")
            return False

        target_path = self.current_path
        if target_path is None:
            base_name = self.controls.title_lbl.text().strip() or "shared_data"
            default_name = f"{Path(base_name).stem or 'shared_data'}.csv"
            picked, _ = QtWidgets.QFileDialog.getSaveFileName(
                self,
                "Save File",
                str(Path.home() / default_name),
                "Data Files (*.csv *.xls *.xlsx *.jsonl)",
            )
            if not picked:
                return False
            target_path = Path(picked)
            if target_path.suffix.lower() not in [".csv", ".xls", ".xlsx", ".jsonl"]:
                target_path = target_path.with_suffix(".csv")
            self.current_path = target_path

        try:
            self._write_dataframe(target_path, self.model)
            self.model.is_dirty = False
            self.status.showMessage(f"Saved to {target_path}")
            return True
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to save:\n{e}")
            return False
    
    def save_copy(self):
        if not self.model:
            return
        prefer_xlsx = (
            self.model.source_format == ".xlsx"
            or (self.model.source_format in [".csv", ".jsonl"] and self.model.has_cell_styles())
        )
        default_path = self.current_path or (Path.home() / "data.csv")
        if prefer_xlsx:
            default_path = default_path.with_suffix(".xlsx")
            file_filter = "Excel (*.xlsx);;CSV (*.csv);;JSON Lines (*.jsonl)"
            initial_filter = "Excel (*.xlsx)"
        else:
            file_filter = "CSV (*.csv);;Excel (*.xlsx);;JSON Lines (*.jsonl)"
            initial_filter = "CSV (*.csv)"

        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save Copy As",
            str(default_path),
            file_filter,
            initial_filter,
        )
        if not path:
            return
        try:
            self._write_dataframe(Path(path), self.model)
            self.status.showMessage(f"Saved copy to {path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to save copy:\n{e}")

    def _write_dataframe(self, path: Path, model: DataFrameModel):
        df = model.df
        suf = path.suffix.lower()
        if suf == ".csv":
            df.to_csv(path, index=False)
        elif suf == ".xlsx":
            base_font = copy(DEFAULT_EXPORT_XLSX_FONT) if DEFAULT_EXPORT_XLSX_FONT is not None else model.xlsx_base_font
            write_xlsx_with_styles(
                path,
                df,
                model.cell_styles(),
                model.xlsx_sheet_name,
                base_font=base_font,
            )
        elif suf == ".xls":
            df.to_excel(path, index=False)
        elif suf == ".jsonl":
            df.to_json(path, orient="records", lines=True, force_ascii=False)
        else:
            df.to_csv(path, index=False)

    def on_search_changed(self, text: str):
        if not self.model:
            return
        self.model.set_search_term(text)
        self.current_match_pos = -1
        self.replace_container.setVisible(bool(text.strip()))
        self._update_replace_buttons()

    def _update_replace_buttons(self):
        has_repl = bool(self.replace_edit.text().strip())
        self.btn_replace_next.setEnabled(has_repl)
        self.btn_replace_all.setEnabled(has_repl)
        if self.model:
            self.model.replace_term = self.replace_edit.text()

    def on_replace_next(self):
        if not self.model:
            return
        idx = self.model.replace_once(self.current_match_index())
        if idx:
            self.current_match_pos = self.model._matches.index(idx)
            
            # Smooth scroll to replaced cell
            row = idx.row()
            v_header = self.table.verticalHeader()
            row_top = v_header.sectionPosition(row)
            row_height = v_header.sectionSize(row)
            viewport_height = self.table.viewport().height()
            
            target_v_scroll = row_top - (viewport_height - row_height) // 2
            
            v_sb = self.table.verticalScrollBar()
            target_v_scroll = max(0, min(target_v_scroll, v_sb.maximum()))
            
            v_sb.setValue(target_v_scroll)
            
            self.table.setCurrentIndex(idx)
            self._on_matches_changed(self.model.total_matches())

    def on_replace_all(self):
        if not self.model:
            return
        n = self.model.replace_all()
        QtWidgets.QMessageBox.information(self, "Replace All", f"Replaced in {n} cell(s).")

    def current_match_index(self) -> t.Optional[QtCore.QModelIndex]:
        if self.current_match_pos >= 0 and self.model:
            return self.model.match_at(self.current_match_pos)
        return None

    def _on_matches_changed(self, total: int):
        current = self.current_match_pos + 1 if total > 0 and self.current_match_pos >= 0 else 0
        self.match_label.setText(f"Matching cells: {current} / {total}")
        self.btn_prev.setEnabled(total > 0)
        self.btn_next.setEnabled(total > 0)

    def go_next_match(self):
        if not self.model or self.model.total_matches() == 0:
            return
        self.current_match_pos = (self.current_match_pos + 1) % self.model.total_matches()
        self._scroll_to_current()

    def go_prev_match(self):
        if not self.model or self.model.total_matches() == 0:
            return
        self.current_match_pos = (self.current_match_pos - 1) % self.model.total_matches()
        self._scroll_to_current()

    def eventFilter(self, obj, event):
        """Capture Shift+Enter in search box for backwards navigation"""
        if obj == self.search_edit and event.type() == QtCore.QEvent.Type.KeyPress:
            if event.key() in (QtCore.Qt.Key.Key_Return, QtCore.Qt.Key.Key_Enter):
                if event.modifiers() & QtCore.Qt.KeyboardModifier.ShiftModifier:
                    # Shift+Enter: go to previous match
                    self.go_prev_match()
                    return True  # Event handled
        return super().eventFilter(obj, event)

    def _scroll_to_current(self):
        idx = self.model.match_at(self.current_match_pos)
        if idx:
            row = idx.row()
            col = idx.column()
            
            # === VERTICAL SCROLLING (Animated) ===
            v_header = self.table.verticalHeader()
            row_top = v_header.sectionPosition(row)
            row_height = v_header.sectionSize(row)
            viewport_height = self.table.viewport().height()
            
            target_v_scroll = row_top - (viewport_height - row_height) // 2
            v_sb = self.table.verticalScrollBar()
            target_v_scroll = max(0, min(target_v_scroll, v_sb.maximum()))
            
            # Use animated setValue (this triggers AnimatedScrollBar's animation)
            v_sb.setValue(int(target_v_scroll))
            
            # === HORIZONTAL SCROLLING (Animated) ===
            h_header = self.table.horizontalHeader()
            col_left = h_header.sectionPosition(col)
            col_width = h_header.sectionSize(col)
            viewport_width = self.table.viewport().width()
            
            target_h_scroll = col_left - (viewport_width - col_width) // 2
            h_sb = self.table.horizontalScrollBar()
            target_h_scroll = max(0, min(target_h_scroll, h_sb.maximum()))
            
            # Use animated setValue for horizontal too
            h_sb.setValue(int(target_h_scroll))
            
            self.table.setCurrentIndex(idx)
            self._on_matches_changed(self.model.total_matches())

    def keyPressEvent(self, event: QtGui.QKeyEvent):
        # Existing Enter key logic
        if event.key() in (QtCore.Qt.Key.Key_Return, QtCore.Qt.Key.Key_Enter):
            if event.modifiers() & QtCore.Qt.KeyboardModifier.ShiftModifier:
                self.go_prev_match()
            else:
                self.go_next_match()
            return
        
        # === NEW: Handle Delete Key ===
        if event.key() == QtCore.Qt.Key.Key_Delete:
            self._handle_delete_key()
            return
            
        super().keyPressEvent(event)
        
    def _handle_delete_key(self):
            """Handles deletion of columns, rows, or cell contents based on selection."""
            if not self.model:
                return

            sel_cols = self.table.selectionModel().selectedColumns()
            if sel_cols:
                col_indices = [c.column() for c in sel_cols]
                if self._remove_columns(col_indices):
                    return

            sel_rows = self.table.selectionModel().selectedRows()
            if sel_rows:
                row_indices = [r.row() for r in sel_rows]
                if self._remove_rows(row_indices):
                    return

            if self.table.selectionModel().hasSelection():
                self.model.layoutAboutToBeChanged.emit()
                for idx in self.table.selectedIndexes():
                    if (
                        idx.isValid()
                        and idx.row() < len(self.model._df.index)
                        and idx.column() < len(self.model._df.columns)
                    ):
                        self.model._df.iat[idx.row(), idx.column()] = ""
                self.model.is_dirty = True
                self.model.layoutChanged.emit()
            
    def autosize_columns(self, force=False):
        if not self.model:
            return
        
        # Cache metrics and calculate once
        font = self.table.font()
        metrics = QtGui.QFontMetrics(font)
        avg_char_width = metrics.averageCharWidth()
        min_w, max_w = 80, 650
        
        # Sample size: 200 rows is enough for most cases
        SAMPLE_SIZE = 200
        df_sample = self.model.df.head(SAMPLE_SIZE) if len(self.model.df) > SAMPLE_SIZE else self.model.df
        
        for c in range(len(self.model.df.columns)):
            header = str(self.model.df.columns[c])
            col = df_sample.iloc[:, c]
            
            # Initialize avg_len with a default value
            avg_len = 8  # ← FIX: Define it here
            
            # Fast path: numeric columns need less width
            if pd.api.types.is_numeric_dtype(col):
                # Find max numeric width
                max_val = col.dropna().abs().max() if not col.dropna().empty else 0
                avg_len = len(f"{max_val:.2f}") if pd.notna(max_val) else 8  # ← Assign here
            else:
                # Use vectorized string length calculation
                sample = col.dropna().astype(str)
                if not sample.empty:
                    avg_len = sample.str.len().mean()
                    avg_len = min(avg_len, 100)  # Cap extremely long text
                else:
                    avg_len = 8  # ← Fallback here
            
            # Now avg_len is always defined
            is_id_col = "id" in header.strip().lower() and avg_len <= 16
            base = 8 if is_id_col else int(avg_len)
            w = int(avg_char_width * (base + len(header) * 0.15)) + 28
            w = max(min_w, min(max_w, w))
            
            # Set width without triggering multiple signals
            self.table.horizontalHeader().resizeSection(c, w)
    
    # === NEW: Column resize handler (debounced) ===
    def _on_column_resized(self, logical_index: int, old_size: int, new_size: int):
        """Triggered when any column is resized. Starts debounced row update."""
        # Ignore tiny adjustments (less than 5 pixels) to avoid micro-calculations
        if abs(new_size - old_size) > 5:
            self._columns_resized.add(logical_index)
            self._row_resize_timer.start()

    # === NEW: Efficient row height updater for visible area ===
    def _resize_visible_rows(self):
        """Recalculates row heights for visible area based on resized columns."""
        if not self.model or not self._columns_resized:
            return
        
        # Get viewport boundaries for visible rows
        viewport = self.table.viewport()
        top_row = self.table.rowAt(0)
        bottom_row = self.table.rowAt(viewport.height())
        
        # Handle edge cases
        if top_row == -1: 
            top_row = 0
        if bottom_row == -1: 
            bottom_row = self.model.rowCount() - 1
        
        # Add buffer rows for smoother scrolling (25 rows above/below)
        BUFFER_ROWS = 25
        top_row = max(0, top_row - BUFFER_ROWS)
        bottom_row = min(self.model.rowCount() - 1, bottom_row + BUFFER_ROWS)
        
        # Calculate new heights only for rows in visible+buffer zone
        for row in range(top_row, bottom_row + 1):
            self._resize_row_with_padding(row)
        
        # Clear the tracking set
        self._columns_resized.clear()
        
        # Background full update for very large tables (>1000 rows)
        if self.model.rowCount() > 1000:
            QtCore.QTimer.singleShot(800, self._background_resize_rows)
            
    def _resize_row_with_padding(self, row: int):
        """
        Resize row to fit content and add a small amount of extra vertical space
        so wrapped text is easier to scan.
        """
        # Resize to fit content first
        self.table.resizeRowToContents(row)

        extra_padding = 3
        current_height = self.table.rowHeight(row)
        new_height = current_height + extra_padding
        self.table.setRowHeight(row, new_height)
        
    # === NEW: Background processing for off-screen rows ===
    def _background_resize_rows(self):
        """Low-priority update for rows outside viewport (runs after idle)."""
        if not self.model:
            return
        
        total_rows = self.model.rowCount()
        
        # Process in batches of 50 to maintain UI responsiveness
        BATCH_SIZE = 50
        delay = 15  # milliseconds between batches
        
        # Use closure to maintain state
        def create_batch_processor():
            current_row = 0
            def process_next_batch():
                nonlocal current_row
                if not self.model:  # Safety check
                    return
                    
                end_row = min(current_row + BATCH_SIZE, total_rows)
                for row in range(current_row, end_row):
                    self._resize_row_with_padding(row)
                
                current_row = end_row
                if current_row < total_rows:
                    QtCore.QTimer.singleShot(delay, process_next_batch)
            
            return process_next_batch
        
        # Start processing from first row
        processor = create_batch_processor()
        processor()

    # === NEW: Handle cell data changes ===
    def _on_cell_data_changed(self, top_left, bottom_right, roles):
        """Recalculate row height when cell data changes via editing."""
        if not roles or QtCore.Qt.ItemDataRole.DisplayRole in roles:
            # Recalculate just the changed rows
            for row in range(top_left.row(), bottom_right.row() + 1):
                self._resize_row_with_padding(row)
        self._update_edit_meta_indicator()

    # === NEW: Handle scrolling for on-demand row sizing ===
    def _on_scroll(self, value):
        """Precalculate rows as they come into view - optimized version"""
        if not self.model:
            return
        
        # More aggressive throttling
        if abs(value - self._last_scroll_value) < 150:
            return
        
        self._last_scroll_value = value
        
        # Get viewport boundaries more efficiently
        viewport_height = self.table.viewport().height()
        if viewport_height <= 0:
            return
        
        top_row = self.table.rowAt(0)
        bottom_row = self.table.rowAt(viewport_height)
        
        # Validate row indices
        if top_row == -1:
            top_row = max(0, int(value / self.table.verticalHeader().defaultSectionSize()) - 10)
        if bottom_row == -1:
            bottom_row = min(self.model.rowCount() - 1, top_row + 100)
        
        # Add smaller buffer for better performance
        BUFFER_ROWS = 20
        top_row = max(0, top_row - BUFFER_ROWS)
        bottom_row = min(self.model.rowCount() - 1, bottom_row + BUFFER_ROWS)
        
        # Check if rows already have reasonable height
        font_metrics = QtGui.QFontMetrics(self.table.font())
        min_expected_height = font_metrics.lineSpacing() + TABLE_VERTICAL_TEXT_PADDING
        
        # Process only rows that need it
        for row in range(top_row, bottom_row + 1):
            if row >= len(self.model.df.index):
                continue
            if self.table.rowHeight(row) < min_expected_height:
                # Double-check if this row actually needs resizing
                needs_resize = False
                for c in range(self.model.columnCount() - 1):
                    val = self.model.df.iat[row, c]
                    if pd.notna(val):
                        s = str(val)
                        if '\n' in s or len(s) > 100:  # Only resize if multiline or very long
                            needs_resize = True
                            break
                
                if needs_resize:
                    self._resize_row_with_padding(row)
                    
    # === NEW: Full recalculation on initial load ===
    def _autosize_all_rows(self):
        """
        Optimized: Sets a default height for all rows instantly, then uses 
        Pandas vectorization to find ONLY the complex rows that need 
        calculation.
        """
        if not self.model:
            return
        
        row_count = self.model.rowCount()
        if row_count == 0:
            return

        # 1. Set global default height (Instant)
        # This handles 90-99% of rows without ANY calculation overhead
        font = self.table.font()
        metrics = QtGui.QFontMetrics(font)
        default_height = metrics.lineSpacing() + TABLE_VERTICAL_TEXT_PADDING
        self.table.verticalHeader().setDefaultSectionSize(default_height)
        
        # 2. Identify "Complex" rows using Vectorized Pandas (Fast)
        # We only care about columns that are 'object' type (text)
        text_cols = self.model.df.select_dtypes(include=['object']).columns
        
        complex_rows_set = set()
        
        if len(text_cols) > 0:
            # Show busy cursor if file is huge
            if row_count > 10000:
                QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)

            try:
                # Check each text column for newlines or excessive length
                # We do this column by column to save memory
                for col in text_cols:
                    series = self.model.df[col].dropna().astype(str)
                    
                    # Criteria: Contains newline OR is longer than 100 chars
                    # We get the boolean mask and then the indices
                    mask = (series.str.contains('\n')) | (series.str.len() > 100)
                    
                    # Add indices of complex rows to our set
                    complex_rows_set.update(series[mask].index.tolist())
            finally:
                 if row_count > 10000:
                    QtWidgets.QApplication.restoreOverrideCursor()

        # Convert to a sorted list for processing
        self._pending_resize_rows = sorted(list(complex_rows_set))
        total_complex = len(self._pending_resize_rows)
        
        # If very few complex rows, just do it now and finish
        if total_complex < 50:
            for row in self._pending_resize_rows:
                self._resize_row_with_padding(row)
            return

        # 3. Setup Background Processing for the Complex Rows ONLY
        self._show_background_indicator("Autosizing rows", total_complex, 0)
        
        # Process first batch immediately to give instant feedback
        BATCH_SIZE = 20
        initial_batch = self._pending_resize_rows[:BATCH_SIZE]
        self._pending_resize_rows = self._pending_resize_rows[BATCH_SIZE:]
        
        for row in initial_batch:
            self._resize_row_with_padding(row)
            
        # Schedule the rest
        if self._pending_resize_rows:
            QtCore.QTimer.singleShot(50, self._background_resize_remaining)
        else:
            self._hide_background_indicator()

    def _background_resize_remaining(self):
        """
        Optimized: Process the queue of specific rows that need resizing.
        Does NOT iterate through the whole table, only the necessary rows.
        """
        # Safety check: ensure the queue exists and model is loaded
        if not self.model or not hasattr(self, '_pending_resize_rows') or not self._pending_resize_rows:
            self._hide_background_indicator()
            return
            
        # Process in batches
        BATCH_SIZE = 50 
        
        # Take the next batch of indices from the queue
        current_batch = self._pending_resize_rows[:BATCH_SIZE]
        self._pending_resize_rows = self._pending_resize_rows[BATCH_SIZE:]
        
        # Resize only these specific rows
        for row in current_batch:
            self._resize_row_with_padding(row)
            
        # Update progress based on how many are LEFT
        if self.bg_total_rows > 0:
            remaining = len(self._pending_resize_rows)
            done = self.bg_total_rows - remaining
            self.bg_current_row = done
            self._update_bg_indicator()
        
        # If there are still rows left, schedule the next batch
        if self._pending_resize_rows:
            # Short delay (10ms) to keep UI responsive
            QtCore.QTimer.singleShot(10, self._background_resize_remaining)
        else:
            self._hide_background_indicator()
            self.status.showMessage(f"Finished formatting.")    
        
    def toggle_mode(self, checked: bool):
        if self.model:
            self.model._is_dark = checked
        if checked:
            apply_dark_palette(QtWidgets.QApplication.instance())
            self.mode_toggle.setText("Dark")
            self.table.setStyleSheet("QTableView { gridline-color: rgb(80, 80, 80); }")
            self._apply_scrollbar_style()
            # Dark mode style for case-sensitive button
            self.case_sensitive_btn.setStyleSheet("""
                QToolButton:checked {
                    background-color: rgb(60, 100, 160);
                    border: 1px solid rgb(80, 130, 200);
                }
            """)
        else:
            apply_light_palette(QtWidgets.QApplication.instance())
            self.mode_toggle.setText("Light")
            # Light mode style for case-sensitive button
            self.case_sensitive_btn.setStyleSheet("""
                QToolButton:checked {
                    background-color: rgb(0, 122, 204);
                    border: 1px solid rgb(0, 100, 180);
                }
            """)
            light_scrollbar_style = """
                QTableView { gridline-color: rgb(0, 0, 0); }
                QScrollBar:vertical {
                    border: none;
                    background: rgb(240, 240, 240);
                    width: 14px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background: rgb(150, 150, 150);
                    min-height: 30px;
                    border-radius: 7px;
                }
                QScrollBar::handle:vertical:hover {
                    background: rgb(120, 120, 120);
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }
                
                QScrollBar:horizontal {
                    border: none;
                    background: rgb(240, 240, 240);
                    height: 14px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background: rgb(150, 150, 150);
                    min-width: 30px;
                    border-radius: 7px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: rgb(120, 120, 120);
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: none;
                }
            """
            self.table.setStyleSheet(light_scrollbar_style)
        self.table.viewport().update()
                
    def _handle_plus_column_click(self, index: QtCore.QModelIndex):
        """Handle clicks on '+' row/column sentinels to add row/column."""
        if not self.model:
            return

        if index.row() == self.model.data_row_count():
            self._add_new_row()
            return

        if index.column() == self.model.data_column_count():
            self._prompt_add_column()

    def _show_cell_qa_menu(self, pos: QtCore.QPoint):
        if not self.model:
            return

        index = self.table.indexAt(pos)
        if not index.isValid():
            return
        if index.row() >= len(self.model.df.index):
            return
        if index.column() >= len(self.model.df.columns):
            return

        self.table.setCurrentIndex(index)
        menu = QtWidgets.QMenu(self)
        mark_pass = menu.addAction("✔️ All Good")
        mark_fail = menu.addAction("❌ Issue")
        menu.addSeparator()
        clear_mark = menu.addAction("Clear mark")

        chosen = menu.exec(self.table.viewport().mapToGlobal(pos))
        if chosen == mark_pass:
            self._set_cell_qa_mark(index, "pass")
        elif chosen == mark_fail:
            self._set_cell_qa_mark(index, "fail")
        elif chosen == clear_mark:
            self._set_cell_qa_mark(index, None)

    def _set_cell_qa_mark(self, index: QtCore.QModelIndex, mark: t.Optional[str]):
        if not self.model or not index.isValid():
            return
        if index.column() >= len(self.model.df.columns):
            return

        if self.model.set_cell_qa_mark(index.row(), index.column(), mark):
            row = index.row() + 1
            col = index.column() + 1
            if mark == "pass":
                self.status.showMessage(f"Marked R{row}C{col} as ✓", 2000)
            elif mark == "fail":
                self.status.showMessage(f"Marked R{row}C{col} as ✗", 2000)
            else:
                self.status.showMessage(f"Cleared mark on R{row}C{col}", 2000)
            self._update_edit_meta_indicator()

    def _rename_column(self, logical_index: int):
        """Rename a column via double-click on header"""
        if not self.model or logical_index >= len(self.model.df.columns):
            return
        
        current_name = self.model.df.columns[logical_index]
        new_name, ok = QtWidgets.QInputDialog.getText(
            self, "Rename Column", f"Rename '{current_name}' to:", text=current_name
        )
        if ok and new_name.strip() and new_name != current_name:
            self.model.df.rename(columns={current_name: new_name}, inplace=True)
            self.model.is_dirty = True
            self.model.headerDataChanged.emit(
                QtCore.Qt.Orientation.Horizontal, logical_index, logical_index
            )
            self.status.showMessage(f"Renamed column to: {new_name}")
        
    def update_dirty_indicator(self, is_dirty: bool = None):
        """Update the unsaved changes indicator based on model state"""
        if not self.model:
            self.dirty_indicator.hide()
            self.edit_meta_indicator.hide()
            return
        
        if is_dirty is None:
            is_dirty = self.model.is_dirty
        
        if is_dirty:
            self.dirty_indicator.setText("⚠️ Unsaved changes")
            self.dirty_indicator.show()
        else:
            self.dirty_indicator.hide()
            
class DiffTextEdit(QtWidgets.QTextEdit):
    def __init__(self):
        super().__init__()
        self.setAcceptRichText(False)
        self.setLineWrapMode(QtWidgets.QTextEdit.LineWrapMode.WidgetWidth)
        self._unescaping = False
        self.textChanged.connect(self._auto_unescape)

    def _auto_unescape(self):
        """Convert literal \n, \t, \\, etc. into real control characters."""
        if self._unescaping:
            return
            
        text = self.toPlainText()
        if '\\' not in text:  # Fast exit if nothing to do
            return
            
        # Only convert these specific escape sequences
        pattern = re.compile(r'\\([ntr\\"\'])')
        escape_map = {'n': '\n', 't': '\t', 'r': '\r', '\\': '\\', '"': '"', "'": "'"}
        
        unescaped = pattern.sub(lambda m: escape_map.get(m.group(1), m.group(0)), text)
        
        if unescaped != text:
            self._unescaping = True
            cursor = self.textCursor()
            scroll_pos = self.verticalScrollBar().value()
            self.setPlainText(unescaped)
            self.setTextCursor(cursor)
            self.verticalScrollBar().setValue(scroll_pos)
            self._unescaping = False

class DiffPage(QtWidgets.QWidget):
    backRequested = QtCore.pyqtSignal()

    def __init__(self, app_ref: 'MainWindow'):
        super().__init__()
        self.app_ref = app_ref

        outer = QtWidgets.QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        # Top bar
        self.controls = WindowControls("Compare Texts")
        self.controls.share_btn.hide()
        self.controls.share_count_lbl.hide()
        self.controls.sync_title_balance()
        self.controls.backRequested.connect(self.backRequested.emit)
        self.controls.minimizeRequested.connect(lambda: self.window().showMinimized())
        self.controls.maximizeRestoreRequested.connect(self._toggle_max_restore)
        self.controls.closeRequested.connect(lambda: QtWidgets.QApplication.instance().quit())


        # Small toolbar with live stats
        toolbar = QtWidgets.QToolBar()
        self.similarity_lbl = QtWidgets.QLabel("Similarity: —")
        self.stats_lbl = QtWidgets.QLabel("Δ: —")
        toolbar.addWidget(self.similarity_lbl)
        toolbar.addWidget(self.stats_lbl)

        # Editors side by side
        editors = QtWidgets.QWidget()
        hl = QtWidgets.QHBoxLayout(editors)
        hl.setContentsMargins(8, 8, 8, 8)

        self.left_edit = DiffTextEdit()
        self.right_edit = DiffTextEdit()
        for e, ph in [(self.left_edit, "Left text (original)…"),
                      (self.right_edit, "Right text (modified)…")]:
            e.setPlaceholderText(ph)
            e.setLineWrapMode(QtWidgets.QTextEdit.LineWrapMode.WidgetWidth)
            e.textChanged.connect(self.schedule_compare)
        hl.addWidget(self.left_edit)
        hl.addWidget(self.right_edit)

        outer.addWidget(self.controls)
        outer.addWidget(toolbar)
        outer.addWidget(editors)

        # Enable undo and redo functionality
        self.left_edit.setUndoRedoEnabled(True)
        self.right_edit.setUndoRedoEnabled(True)

        # Store base font size for zooming
        self.base_font_size = self.left_edit.font().pointSize()
        self.zoom_factor = 1.0

        # Timer for debounced comparison
        self.compare_timer = QtCore.QTimer()
        self.compare_timer.setSingleShot(True)
        self.compare_timer.setInterval(100)
        self.compare_timer.timeout.connect(self.run_compare)

        # Set up keyboard shortcuts for undo/redo
        self.set_up_shortcuts()

        # Selection synchronization
        self._syncing_selection = False
        self.left_edit.cursorPositionChanged.connect(lambda: self._sync_selection(self.left_edit, self.right_edit))
        self.right_edit.cursorPositionChanged.connect(lambda: self._sync_selection(self.right_edit, self.left_edit))

    def _sync_selection(self, source_edit: QtWidgets.QTextEdit, target_edit: QtWidgets.QTextEdit):
        """Synchronize text selection using Regex for multi-line and forced cleanup for stuck highlights."""
        if self._syncing_selection:
            return

        self._syncing_selection = True
        try:
            # --- FIX 1: STUCK SELECTION ---
            # Immediately clear highlights on BOTH sides. 
            # This ensures that if you just clicked to deselect (hasSelection is False),
            # the old artifacts are removed immediately.
            source_edit.setExtraSelections([])
            target_edit.setExtraSelections([])

            cursor = source_edit.textCursor()
            if not cursor.hasSelection():
                return

            text = cursor.selectedText()
            if not text.strip():
                return

            # Detect if selection is in diff-only region (red/green background)
            start_cursor = QtGui.QTextCursor(cursor)
            start_cursor.setPosition(cursor.selectionStart())
            char_format = start_cursor.charFormat()
            bg_color = char_format.background().color()

            # Check strictly for the diff colors defined in run_compare
            is_red = bg_color.red() > 150 and bg_color.green() < 80 and bg_color.blue() < 80
            is_green = bg_color.green() > 120 and bg_color.red() < 80 and bg_color.blue() < 80

            if is_red or is_green:
                return

            # --- FIX 2: MULTI-LINE SELECTION ---
            # Split selection into words and create a Regex pattern.
            # This finds the words regardless of whether the gap is a space, \n, or \u2029
            parts = text.split()
            if not parts:
                return
                
            # Escape parts to handle special chars like brackets or dots in the text
            safe_parts = [QtCore.QRegularExpression.escape(p) for p in parts]
            
            # Join words with \s+ (matches any whitespace sequence including newlines)
            pattern = r"\s+".join(safe_parts)
            regex = QtCore.QRegularExpression(pattern)
            
            target_doc = target_edit.document()
            
            # Determine search direction
            sel_start = cursor.selectionStart()
            source_mid = source_edit.document().characterCount() // 2

            if sel_start < source_mid:
                find_cursor = target_doc.find(regex, 0)
            else:
                text_length = target_doc.characterCount()
                find_cursor = target_doc.find(regex, text_length - 1,
                                              QtGui.QTextDocument.FindFlag.FindBackward)

            if find_cursor.isNull():
                return

            # Scroll alignment
            source_rect = source_edit.cursorRect(cursor)
            source_viewport_height = source_edit.viewport().height()

            if source_viewport_height > 0:
                source_rel_y = source_rect.center().y() / source_viewport_height
                source_rel_y = max(0.0, min(1.0, source_rel_y))

                target_edit.setTextCursor(find_cursor)
                target_rect = target_edit.cursorRect(find_cursor)
                target_viewport_height = target_edit.viewport().height()

                if target_viewport_height > 0:
                    target_scrollbar = target_edit.verticalScrollBar()
                    current_scroll = target_scrollbar.value()
                    desired_y = source_rel_y * target_viewport_height
                    delta_y = target_rect.top() - desired_y
                    new_scroll = current_scroll + delta_y
                    new_scroll = max(0, min(new_scroll, target_scrollbar.maximum()))
                    target_scrollbar.setValue(int(new_scroll))

            # Apply blue highlight
            selection = QtWidgets.QTextEdit.ExtraSelection()
            selection.cursor = find_cursor
            selection.format.setBackground(QtGui.QColor(0, 150, 255, 100))
            selection.format.setProperty(QtGui.QTextFormat.Property.FullWidthSelection, False)

            target_edit.setExtraSelections([selection])
            
        finally:
            # Ensure flag is reset even if we returned early
            self._syncing_selection = False
    def _clear_selection(self, edit: QtWidgets.QTextEdit):
        """Clear extra selection highlights in the given editor"""
        if self._syncing_selection:
            return
        self._syncing_selection = True
        edit.setExtraSelections([])
        edit.viewport().update()
        self._syncing_selection = False
        
    def set_up_shortcuts(self):
        QtGui.QShortcut(QtGui.QKeySequence.StandardKey.Undo, self, self.undo_text)
        QtGui.QShortcut(QtGui.QKeySequence.StandardKey.Redo, self, self.redo_text)

    def undo_text(self):
        focused = QtWidgets.QApplication.focusWidget()
        if focused in (self.left_edit, self.right_edit):
            if focused.isUndoAvailable():
                focused.undo()

    def redo_text(self):
        focused = QtWidgets.QApplication.focusWidget()
        if focused in (self.left_edit, self.right_edit):
            if focused.isRedoAvailable():
                focused.redo()

    def schedule_compare(self):
        """Debounce comparison to avoid running on every keystroke"""
        self.compare_timer.start()

    def wheelEvent(self, event: QtGui.QWheelEvent):
        if event.modifiers() == QtCore.Qt.KeyboardModifier.ControlModifier:
            delta = event.angleDelta().y() / 120
            
            self.zoom_factor += delta * 0.1
            self.zoom_factor = max(0.5, min(3.0, self.zoom_factor))
            
            new_size = int(self.base_font_size * self.zoom_factor)
            font = self.left_edit.font()
            font.setPointSize(new_size)
            self.left_edit.setFont(font)
            self.right_edit.setFont(font)
            
            event.accept()
        else:
            super().wheelEvent(event)

    def run_compare(self):
        from html import escape as esc
        import difflib
        import re

        left = self.left_edit.toPlainText()
        right = self.right_edit.toPlainText()
        
        if not left and not right:
            self.similarity_lbl.setText("Similarity: —")
            self.stats_lbl.setText("Δ: —")
            return

        left_cursor = self.left_edit.textCursor()
        right_cursor = self.right_edit.textCursor()
        left_scroll = self.left_edit.verticalScrollBar().value()
        right_scroll = self.right_edit.verticalScrollBar().value()

        def words(s: str):
            return re.findall(r"\s+|[a-zA-Z0-9]+|[^a-zA-Z0-9\s]", s)

        l_words, r_words = words(left), words(right)
        sm = difflib.SequenceMatcher(a=l_words, b=r_words)
        opcodes = sm.get_opcodes()

        left_html_parts, right_html_parts = [], []
        dels = adds = 0

        for tag, i1, i2, j1, j2 in opcodes:
            l_chunk = "".join(l_words[i1:i2])
            r_chunk = "".join(r_words[j1:j2])
            if tag == "equal":
                left_html_parts.append(esc(l_chunk))
                right_html_parts.append(esc(r_chunk))
            elif tag == "delete":
                left_html_parts.append(f'<span style="background-color: rgba(255,0,0,0.35);">{esc(l_chunk)}</span>')
                dels += 1
            elif tag == "insert":
                right_html_parts.append(f'<span style="background-color: rgba(0,200,0,0.35);">{esc(r_chunk)}</span>')
                adds += 1
            elif tag == "replace":
                left_html_parts.append(f'<span style="background-color: rgba(255,0,0,0.35);">{esc(l_chunk)}</span>')
                right_html_parts.append(f'<span style="background-color: rgba(0,200,0,0.35);">{esc(r_chunk)}</span>')
                dels += 1
                adds += 1

        left_html = "<div style='font-family: Consolas, monospace; white-space: pre-wrap;'>" + "".join(left_html_parts) + "</div>"
        right_html = "<div style='font-family: Consolas, monospace; white-space: pre-wrap;'>" + "".join(right_html_parts) + "</div>"

        self.left_edit.blockSignals(True)
        self.right_edit.blockSignals(True)
        
        self.left_edit.setHtml(left_html)
        self.right_edit.setHtml(right_html)
        
        self.left_edit.setTextCursor(left_cursor)
        self.right_edit.setTextCursor(right_cursor)
        
        self.left_edit.verticalScrollBar().setValue(left_scroll)
        self.right_edit.verticalScrollBar().setValue(right_scroll)
        
        self.left_edit.blockSignals(False)
        self.right_edit.blockSignals(False)

        sim = sm.ratio()
        self.similarity_lbl.setText(f"Similarity: {sim*100:.1f}%")
        self.stats_lbl.setText(f"Δ: removals (left) {dels} | additions (right) {adds}")

    def _toggle_max_restore(self):
        if self.window().isMaximized() or self.window().isFullScreen():
            self.window().showNormal()
        else:
            self.window().showMaximized()

class StartPage(QtWidgets.QWidget):
    viewRequested = QtCore.pyqtSignal()
    compareRequested = QtCore.pyqtSignal()

    def __init__(self):
        super().__init__()
        
        # Main layout with no margins/spacing for the window controls
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Add window controls bar - hide back button on start page
        self.controls = WindowControls(APP_NAME)
        self.controls.back_btn.hide()
        self.controls.share_btn.hide()
        self.controls.share_count_lbl.hide()
        self.controls.sync_title_balance()
        self.controls.minimizeRequested.connect(lambda: self.window().showMinimized())
        self.controls.maximizeRestoreRequested.connect(self._toggle_max_restore)
        self.controls.closeRequested.connect(QtWidgets.QApplication.instance().quit)
        
        # Content container for centered elements
        content = QtWidgets.QWidget()
        content_layout = QtWidgets.QVBoxLayout(content)
        content_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        
        # Welcome label
        lbl = QtWidgets.QLabel("Choose your mode")
        lbl.setStyleSheet("font-size: 20px; font-weight: 600;")
        lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        
        # Buttons container
        btns = QtWidgets.QWidget()
        h = QtWidgets.QHBoxLayout(btns)
        h.setSpacing(20)
        
        view = QtWidgets.QPushButton("View File")
        diff = QtWidgets.QPushButton("Compare Texts")
        
        # Style the buttons
        button_style = """
            QPushButton {
                min-height: 80px;
                font-size: 16px;
                font-weight: 600;
                border-radius: 12px;
                padding: 10px 20px;
                background-color: rgb(45, 45, 45);
                border: 1px solid rgb(60, 60, 60);
            }
            QPushButton:hover {
                background-color: rgba(60, 100, 160, 180);
            }
        """
        view.setStyleSheet(button_style)
        diff.setStyleSheet(button_style)
        
        view.clicked.connect(self.viewRequested.emit)
        diff.clicked.connect(self.compareRequested.emit)
        
        h.addWidget(view)
        h.addWidget(diff)
        
        # Assemble content
        content_layout.addWidget(lbl)
        content_layout.addWidget(btns)
        content_layout.addSpacing(10)
        
        # Add to main layout
        layout.addWidget(self.controls)
        layout.addWidget(content, 1)  # Stretch factor for content

    def _toggle_max_restore(self):
        """Toggle between maximized and normal window state"""
        if self.window().isMaximized():
            self.window().showNormal()
        else:
            self.window().showMaximized()

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.setWindowIcon(QtGui.QIcon())
        self.setWindowFlags(QtCore.Qt.WindowType.FramelessWindowHint)
        central = QtWidgets.QStackedWidget()
        self.setCentralWidget(central)

        self.start = StartPage()
        self.viewer = DataViewerPage(self)
        self.diff = DiffPage(self)

        central.addWidget(self.start)
        central.addWidget(self.viewer)
        central.addWidget(self.diff)
        central.setCurrentWidget(self.viewer)

        # Start directly in viewer mode and keep other modes unreachable.
        self.viewer.backRequested.connect(self.back_to_viewer_dropzone)

        self.resize(COMPACT_WINDOW_WIDTH, COMPACT_WINDOW_HEIGHT)
        self.center_on_screen()

        apply_dark_palette(QtWidgets.QApplication.instance())

        self._create_shortcuts()
        
        self._last_screen = None
        self._screen_change_timer = QtCore.QTimer()
        self._screen_change_timer.setSingleShot(True)
        self._screen_change_timer.setInterval(100)
        self._is_initial_move = True

    def closeEvent(self, event):
        """Intercept close to check for unsaved changes"""
        current_widget = self.centralWidget().currentWidget()
        
        if current_widget == self.viewer and self.viewer.model and self.viewer.model.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self, 
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before closing?",
                (QtWidgets.QMessageBox.StandardButton.Yes | 
                QtWidgets.QMessageBox.StandardButton.No | 
                QtWidgets.QMessageBox.StandardButton.Cancel)
            )

            if reply == QtWidgets.QMessageBox.StandardButton.Cancel:
                event.ignore()
                return
            elif reply == QtWidgets.QMessageBox.StandardButton.Yes:
                if not self.viewer.save_file():
                    event.ignore()
                    return

        self.viewer.shutdown()
        event.accept()
        
    def _create_shortcuts(self):
        QtGui.QShortcut(QtGui.QKeySequence.StandardKey.Find, self, activated=self._focus_search)
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+S"), self, activated=lambda: self.viewer.save_file())
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+O"), self, activated=lambda: self.viewer.open_file_dialog())
        QtGui.QShortcut(QtGui.QKeySequence("Esc"), self, activated=self._escape_behavior)

    def _focus_search(self):
        if self.centralWidget().currentWidget() is self.viewer:
            self.viewer.search_edit.setFocus()

    def _escape_behavior(self):
        if self.isFullScreen():
            self.showNormal()

    def back_to_viewer_dropzone(self):
        """Return to the viewer drag-and-drop screen."""
        if self.centralWidget().currentWidget() != self.viewer:
            return

        if self.viewer.model and self.viewer.model.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before leaving?",
                (QtWidgets.QMessageBox.StandardButton.Yes |
                QtWidgets.QMessageBox.StandardButton.No |
                QtWidgets.QMessageBox.StandardButton.Cancel)
            )

            if reply == QtWidgets.QMessageBox.StandardButton.Cancel:
                return
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                if not self.viewer.save_file():
                    return

        self.viewer.reset_viewer()
        self.centralWidget().setCurrentWidget(self.viewer)
        self.showNormal()
        self.resize(COMPACT_WINDOW_WIDTH, COMPACT_WINDOW_HEIGHT)
        self.center_on_screen()

    def center_on_screen(self):
        # Get the screen that contains the mouse cursor
        cursor_pos = QtGui.QCursor.pos()
        screen = QtWidgets.QApplication.screenAt(cursor_pos)
        
        # Fallback to primary screen if screenAt returns None
        if screen is None:
            screen = QtGui.QGuiApplication.primaryScreen()
        
        screen_geometry = screen.availableGeometry()  # Use availableGeometry to account for taskbars
        
        # Calculate center position
        x = screen_geometry.x() + (screen_geometry.width() - self.width()) // 2
        y = screen_geometry.y() + (screen_geometry.height() - self.height()) // 2
        
        self.move(x, y)

    
    
    def back_to_start(self):
        """Return to start screen with unsaved changes check"""
        # Check if in viewer with unsaved changes
        if self.centralWidget().currentWidget() == self.viewer and self.viewer.model and self.viewer.model.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before leaving?",
                (QtWidgets.QMessageBox.StandardButton.Yes | 
                QtWidgets.QMessageBox.StandardButton.No | 
                QtWidgets.QMessageBox.StandardButton.Cancel)
            )

            if reply == QtWidgets.QMessageBox.StandardButton.Cancel:
                return  # Abort - stay on current screen
            elif reply == QtWidgets.QMessageBox.StandardButton.Yes:
                if not self.viewer.save_file():
                    return  # Abort if save failed
        
        # Clear viewer state (NEW)
        self.viewer.reset_viewer()
        
        # Only proceed if user didn't cancel
        self.centralWidget().setCurrentWidget(self.start)
        self.showNormal()
        self.resize(COMPACT_WINDOW_WIDTH, COMPACT_WINDOW_HEIGHT)
        self.center_on_screen()
        

        
    def moveEvent(self, event):
        """Detect when window moves and check for screen changes"""
        super().moveEvent(event)
        # Trigger screen change check with debounce
        self._screen_change_timer.start()

    def enter_viewer(self):
        self.centralWidget().setCurrentWidget(self.viewer)
        # --- CHANGE: Use maximized instead of full-screen to avoid DPI issues ---
        if self.viewer.model is not None:
            self.showMaximized()  # Better than fullScreen for multi-screen
        else:
            self.showNormal()
            self.resize(COMPACT_WINDOW_WIDTH, COMPACT_WINDOW_HEIGHT)
            self.center_on_screen()

    def enter_diff(self):
        self.centralWidget().setCurrentWidget(self.diff)
        self.showMaximized()  # Better than fullScreen for multi-screen

def main():
    from PyQt6 import QtCore
    
    QtWidgets.QApplication.setHighDpiScaleFactorRoundingPolicy(
        QtCore.Qt.HighDpiScaleFactorRoundingPolicy.Round
    )
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app_font = QtGui.QFont(app.font())
    app_font.setFamily(DEFAULT_DISPLAY_FONT_FAMILY)
    app_font.setPointSize(DEFAULT_DISPLAY_FONT_SIZE)
    app.setFont(app_font)
    win = MainWindow()
    win.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
